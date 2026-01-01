"""
SISTEMA DE INVENTARIO V7.0 HÍBRIDO - MYSQL + UI MEJORADA
==========================================================
PARTE 1 DE 3: Configuración, Logging, Base de Datos y Utilidades

CARACTERÍSTICAS:
- Base de datos MySQL multi-usuario
- Sincronización automática cada 10 segundos
- Detección de conflictos entre equipos
- Hilos para no bloquear UI
- Interfaz responsive
- Export Excel multi-hoja

DEPENDENCIAS:
pip install customtkinter mysql-connector-python pandas openpyxl
"""

import json
import logging
import os
import sys
import threading
import time
import winsound
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox
from typing import Any, Dict, List, Tuple

import customtkinter as ctk
import mysql.connector
import pandas as pd
from mysql.connector import Error, pooling
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================================================================
# CONFIGURACIÓN DE LOGGING
# ============================================================================
LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / f"inventario_{datetime.now().strftime('%Y%m%d')}.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURACIÓN VISUAL
# ============================================================================
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")


class Colors:
    BG = "#1a1a1a"
    CARD = "#2B2B2B"
    ACCENT = "#3B8ED0"
    SUCCESS = "#2CC985"
    DANGER = "#E74C3C"
    WARNING = "#F39C12"
    INFO = "#9B59B6"
    CONSOLE = "#000000"


# ============================================================================
# GESTOR DE CONFIGURACIÓN
# ============================================================================
class ConfigManager:
    CONFIG_FILE = Path("config.json")
    DEFAULT_CONFIG = {
        'database': {
            'host': 'localhost',
            'port': 3306,
            'user': 'root',
            'password': '',
            'database': 'sis_inventario_db',
            'pool_size': 10,
            'pool_name': 'inventario_pool'
        },
        'app': {
            'sync_interval_seconds': 30,
            'max_results_display': 50,
            'enable_audit_log': True,
            'auto_backup': True
        }
    }
    
    @classmethod
    def load(cls) -> Dict[str, Any]:
        try:
            if cls.CONFIG_FILE.exists():
                with open(cls.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    logger.info("Configuración cargada")
                    return config
            else:
                cls.save(cls.DEFAULT_CONFIG)
                logger.warning("Config creado por defecto")
                return cls.DEFAULT_CONFIG
        except Exception as e:
            logger.error(f"Error config: {e}")
            return cls.DEFAULT_CONFIG
    
    @classmethod
    def save(cls, config: Dict[str, Any]) -> bool:
        try:
            with open(cls.CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4)
            return True
        except Exception as e:
            logger.error(f"Error guardando config: {e}")
            return False


# ============================================================================
# GESTOR DE BASE DE DATOS
# ============================================================================
class DBManager:
    def __init__(self):
        self.config = ConfigManager.load()
        self.db_config = self.config['database']
        self.connection_pool = None
        self._initialize_database()
        self._initialize_pool()
    
    def _initialize_database(self):
        try:
            temp_config = {
                'host': self.db_config['host'],
                'port': self.db_config.get('port', 3306),
                'user': self.db_config['user'],
                'password': self.db_config['password']
            }
            
            conn = mysql.connector.connect(**temp_config)
            cursor = conn.cursor()
            cursor.execute(
                f"CREATE DATABASE IF NOT EXISTS {self.db_config['database']} "
                f"CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
            )
            logger.info(f"BD '{self.db_config['database']}' OK")
            cursor.close()
            conn.close()
        except Error as e:
            logger.error(f"Error creando BD: {e}")
            raise
    
    def _initialize_pool(self):
        try:
            pool_config = {
                'host': self.db_config['host'],
                'port': self.db_config.get('port', 3306),
                'user': self.db_config['user'],
                'password': self.db_config['password'],
                'database': self.db_config['database'],
                'pool_size': self.db_config.get('pool_size', 10),
                'pool_name': self.db_config.get('pool_name', 'inventario_pool')
            }
            
            self.connection_pool = pooling.MySQLConnectionPool(**pool_config)
            logger.info("Pool OK")
            self._create_tables()
        except Error as e:
            logger.error(f"Error pool: {e}")
            raise
    
    @contextmanager
    def get_connection(self):
        conn = None
        try:
            conn = self.connection_pool.get_connection()
            yield conn
        except Error as e:
            logger.error(f"Error conexión: {e}")
            if conn:
                conn.rollback()
            raise
        finally:
            if conn and conn.is_connected():
                conn.close()
    
    def _create_tables(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Sesiones
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS sesiones (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    nombre VARCHAR(100) NOT NULL,
                    fecha DATETIME NOT NULL,
                    responsable VARCHAR(100),
                    bodega VARCHAR(50),
                    activo BOOLEAN DEFAULT 1,
                    INDEX idx_activo (activo)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            
            # Items
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS items_corte (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    sesion_id INT NOT NULL,
                    codigo VARCHAR(50) NOT NULL,
                    producto TEXT,
                    linea VARCHAR(100),
                    stock_sistema DECIMAL(10,2) DEFAULT 0,
                    conteo_fisico DECIMAL(10,2) DEFAULT 0,
                    diferencia DECIMAL(10,2) GENERATED ALWAYS AS (conteo_fisico - stock_sistema) VIRTUAL,
                    novedad TEXT,
                    fecha_conteo DATETIME,
                    ultimo_equipo_id INT,
                    FOREIGN KEY (sesion_id) REFERENCES sesiones(id) ON DELETE CASCADE,
                    INDEX idx_sesion_codigo (sesion_id, codigo),
                    INDEX idx_linea (linea)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            
            # Equipos
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS equipos (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    nombre_equipo VARCHAR(100) UNIQUE NOT NULL,
                    integrantes TEXT,
                    activo BOOLEAN DEFAULT 1,
                    fecha_creacion DATETIME DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_activo (activo)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            
            # Verificar si la columna integrantes existe, si no, agregarla
            cursor.execute("""
                SELECT COUNT(*) as existe 
                FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                AND TABLE_NAME = 'equipos' 
                AND COLUMN_NAME = 'integrantes'
            """)
            
            resultado = cursor.fetchone()
            if resultado[0] == 0:
                cursor.execute("ALTER TABLE equipos ADD COLUMN integrantes TEXT AFTER nombre_equipo")
                logger.info("Columna integrantes agregada a tabla equipos")
            
            # Historial
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS historial_movimientos (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    sesion_id INT NOT NULL,
                    item_codigo VARCHAR(50),
                    equipo_id INT,
                    tipo_accion VARCHAR(50),
                    cantidad_anterior DECIMAL(10,2),
                    cantidad_resultante DECIMAL(10,2),
                    fecha_movimiento DATETIME NOT NULL,
                    FOREIGN KEY (sesion_id) REFERENCES sesiones(id) ON DELETE CASCADE,
                    INDEX idx_sesion (sesion_id),
                    INDEX idx_fecha (fecha_movimiento)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            
            conn.commit()
            logger.info("Tablas OK")
    
    def execute_query(self, query: str, params: tuple = None, fetch: bool = False):
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor(dictionary=True) if fetch else conn.cursor()
                cursor.execute(query, params or ())
                
                if fetch:
                    result = cursor.fetchall()
                    cursor.close()
                    return result
                else:
                    conn.commit()
                    last_id = cursor.lastrowid
                    cursor.close()
                    return last_id
        except Error as e:
            logger.error(f"Error query: {e}")
            raise


# ============================================================================
# GESTOR DE BACKUPS
# ============================================================================
class BackupManager:
    BACKUP_DIR = Path("BACKUPS_INVENTARIO")
    
    def __init__(self, db: 'DBManager'):
        self.db = db
        self.BACKUP_DIR.mkdir(exist_ok=True)
    
    def crear_backup(self) -> Tuple[bool, str]:
        """Crea backup completo de la BD"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = self.BACKUP_DIR / f"backup_{timestamp}.sql"
            
            db_config = self.db.db_config
            
            # Construir comando mysqldump
            cmd = (
                f'mysqldump -h {db_config["host"]} '
                f'-P {db_config.get("port", 3306)} '
                f'-u {db_config["user"]} '
            )
            
            if db_config['password']:
                cmd += f'-p"{db_config["password"]}" '
            
            cmd += f'{db_config["database"]} > "{backup_file}"'
            
            # Ejecutar comando
            import subprocess
            result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
            
            if result.returncode == 0 and backup_file.exists():
                size = backup_file.stat().st_size / 1024  # KB
                logger.info(f"Backup creado: {backup_file.name} ({size:.2f} KB)")
                return True, str(backup_file)
            else:
                error_msg = result.stderr if result.stderr else "Error desconocido"
                logger.error(f"Error backup: {error_msg}")
                return False, f"Error: {error_msg}"
        
        except Exception as e:
            logger.error(f"Error creando backup: {e}")
            return False, str(e)
    
    def restaurar_backup(self, backup_path: str) -> Tuple[bool, str]:
        """Restaura BD desde backup"""
        try:
            if not Path(backup_path).exists():
                return False, "Archivo no existe"
            
            db_config = self.db.db_config
            
            # Construir comando mysql
            cmd = (
                f'mysql -h {db_config["host"]} '
                f'-P {db_config.get("port", 3306)} '
                f'-u {db_config["user"]} '
            )
            
            if db_config['password']:
                cmd += f'-p"{db_config["password"]}" '
            
            cmd += f'{db_config["database"]} < "{backup_path}"'
            
            # Ejecutar comando
            import subprocess
            result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
            
            if result.returncode == 0:
                logger.info(f"Backup restaurado: {backup_path}")
                return True, "Backup restaurado exitosamente"
            else:
                error_msg = result.stderr if result.stderr else "Error desconocido"
                logger.error(f"Error restaurando: {error_msg}")
                return False, f"Error: {error_msg}"
        
        except Exception as e:
            logger.error(f"Error restaurando backup: {e}")
            return False, str(e)
    
    def listar_backups(self) -> List[Dict[str, Any]]:
        """Lista todos los backups disponibles"""
        backups = []
        try:
            for file in sorted(self.BACKUP_DIR.glob("*.sql"), reverse=True):
                stat = file.stat()
                backups.append({
                    'nombre': file.name,
                    'path': str(file),
                    'fecha': datetime.fromtimestamp(stat.st_mtime),
                    'tamaño_kb': stat.st_size / 1024
                })
        except Exception as e:
            logger.error(f"Error listando backups: {e}")
        
        return backups
    
    def eliminar_backup(self, backup_path: str) -> bool:
        """Elimina un backup"""
        try:
            Path(backup_path).unlink()
            logger.info(f"Backup eliminado: {backup_path}")
            return True
        except Exception as e:
            logger.error(f"Error eliminando backup: {e}")
            return False


# ============================================================================
# VALIDADORES
# ============================================================================
class InputValidator:
    @staticmethod
    def validate_codigo(codigo: str) -> Tuple[bool, str]:
        codigo = codigo.strip().upper()
        if not codigo:
            return False, "Código vacío"
        if len(codigo) > 50:
            return False, "Código muy largo"
        return True, codigo
    
    @staticmethod
    def validate_cantidad(cantidad_str: str) -> Tuple[bool, float]:
        try:
            cantidad = float(cantidad_str.replace(',', '.'))
            if cantidad < 0:
                return False, "Cantidad negativa"
            if cantidad > 999999:
                return False, "Cantidad muy grande"
            return True, cantidad
        except ValueError:
            return False, "Cantidad inválida"
    
    @staticmethod
    def validate_nombre(nombre: str, max_length: int = 100) -> Tuple[bool, str]:
        nombre = nombre.strip()
        if not nombre:
            return False, "Nombre vacío"
        if len(nombre) > max_length:
            return False, f"Nombre largo (max {max_length})"
        return True, nombre


# ============================================================================
# UTILIDADES
# ============================================================================
class Utils:
    @staticmethod
    def limpiar_codigo(valor):
        """Limpia y normaliza códigos"""
        if isinstance(valor, pd.Series):
            return valor.astype(str).str.strip().str.upper()
        return str(valor).strip().upper()
    
    @staticmethod
    def calcular_stock_desde_excel(excel_path: str) -> Dict[str, float]:
        """Calcula stock sumando múltiples hojas"""
        stock = {}
        hojas = ["CONDI", "MAQUI", "ASCINTEC"]
        
        for hoja in hojas:
            try:
                df = pd.read_excel(excel_path, sheet_name=hoja, dtype=str)
                df.columns = df.columns.str.strip().str.lower()
                
                if 'codproducto' in df.columns and 'sin_stock' in df.columns:
                    for _, row in df.iterrows():
                        codigo = Utils.limpiar_codigo(row['codproducto'])
                        valor = pd.to_numeric(
                            str(row['sin_stock']).replace(",", "."),
                            errors='coerce'
                        )
                        if pd.notnull(valor):
                            # Convertir float64 a float de Python
                            stock[codigo] = float(stock.get(codigo, 0.0) + valor)
            except Exception as e:
                logger.warning(f"Hoja '{hoja}': {e}")
        
        logger.info(f"Stock calculado: {len(stock)} items")
        return stock
    
    @staticmethod
    def cargar_equipos_desde_excel(excel_path: str) -> List[dict]:
        """Carga equipos desde la hoja EQUIPOS del Excel (ID, INTEGRANTES, FECHA DEL EQUIPO)"""
        equipos = []
        try:
            df = pd.read_excel(excel_path, sheet_name="EQUIPOS", dtype=str)
            
            # Buscar columnas (ID, INTEGRANTES, FECHA DEL EQUIPO)
            col_id = None
            col_integrantes = None
            
            for col in df.columns:
                col_upper = str(col).upper().strip()
                if col_upper == 'ID':
                    col_id = col
                elif 'INTEGRANTE' in col_upper:
                    col_integrantes = col
            
            if col_id:
                for idx, row in df.iterrows():
                    num_equipo = str(row[col_id]).strip() if pd.notna(row[col_id]) else None
                    if num_equipo and num_equipo != 'nan':
                        equipo_data = {'numero': num_equipo}
                        if col_integrantes and pd.notna(row[col_integrantes]):
                            equipo_data['integrantes'] = str(row[col_integrantes]).strip().upper()
                        else:
                            equipo_data['integrantes'] = ''
                        equipos.append(equipo_data)
                
                logger.info(f"Equipos cargados desde Excel: {len(equipos)}")
            else:
                logger.warning("No se encontró columna ID en hoja EQUIPOS")
        
        except Exception as e:
            logger.warning(f"No se pudo cargar hoja EQUIPOS: {e}")
        
        return equipos


# ============================================================================
# VENTANA MULTISELECT
# ============================================================================
class VentanaMultiSelect(ctk.CTkToplevel):
    def __init__(self, master, opciones: List[str], callback):
        super().__init__(master)
        self.title("FILTRAR LÍNEAS")
        self.geometry("450x650")
        self.attributes("-topmost", True)
        self.resizable(False, False)
        
        self.callback = callback
        self.vars = []
        
        header = ctk.CTkFrame(self, fg_color=Colors.CARD)
        header.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(header, text="SELECCIONAR LÍNEAS:",
                     font=("Arial", 16, "bold")).pack(pady=10)
        
        btn_frame = ctk.CTkFrame(header, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkButton(btn_frame, text="Todas", width=100,
                      command=self._select_all).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="Ninguna", width=100,
                      command=self._select_none).pack(side="left", padx=5)
        
        self.lbl_count = ctk.CTkLabel(btn_frame, text=f"{len(opciones)} opciones",
                                      text_color="gray")
        self.lbl_count.pack(side="right", padx=10)
        
        self.scroll = ctk.CTkScrollableFrame(self)
        self.scroll.pack(fill="both", expand=True, padx=10, pady=5)
        
        for op in sorted(opciones):
            var = ctk.BooleanVar(value=True)
            cb = ctk.CTkCheckBox(self.scroll, text=op, variable=var,
                                 command=self._update_count)
            cb.pack(anchor="w", pady=3, padx=5)
            self.vars.append((op, var))
        
        ctk.CTkButton(self, text="APLICAR FILTRO", fg_color=Colors.SUCCESS,
                      height=40, font=("Arial", 14, "bold"),
                      command=self._apply).pack(fill="x", padx=20, pady=15)
        
        self._update_count()
    
    def _select_all(self):
        for _, var in self.vars:
            var.set(True)
        self._update_count()
    
    def _select_none(self):
        for _, var in self.vars:
            var.set(False)
        self._update_count()
    
    def _update_count(self):
        sel = sum(1 for _, v in self.vars if v.get())
        self.lbl_count.configure(text=f"{sel}/{len(self.vars)} seleccionadas")
    
    def _apply(self):
        selected = [op for op, var in self.vars if var.get()]
        self.withdraw()
        try:
            self.callback(selected)
        except Exception as e:
            logger.error(f"Error callback filtro: {e}")
        finally:
            self.after(100, self.destroy)

# ============================================================================
# VENTANA GESTIÓN DE BACKUPS
# ============================================================================
class VentanaBackups(ctk.CTkToplevel):
    def __init__(self, master, backup_mgr: BackupManager, callback):
        super().__init__(master)
        self.title("GESTIÓN DE BACKUPS")
        self.geometry("700x600")
        self.attributes("-topmost", True)
        
        self.backup_mgr = backup_mgr
        self.callback = callback
        
        # Header
        header = ctk.CTkFrame(self, fg_color=Colors.INFO)
        header.pack(fill="x")
        
        ctk.CTkLabel(header, text="GESTION DE BACKUPS",
                    font=("Arial", 20, "bold"), text_color="white").pack(pady=15)
        
        # Botones de acción
        btn_frame = ctk.CTkFrame(self, fg_color=Colors.CARD)
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkButton(btn_frame, text="CREAR BACKUP AHORA",
                     fg_color=Colors.SUCCESS, height=40, font=("Arial", 13, "bold"),
                     command=self._crear_backup).pack(side="left", padx=10, pady=10)
        
        ctk.CTkButton(btn_frame, text="REFRESCAR", 
                     fg_color=Colors.ACCENT, height=40, width=120,
                     command=self._load_backups).pack(side="left", padx=10)
        
        # Info
        self.lbl_info = ctk.CTkLabel(self, text="", font=("Arial", 11), 
                                     text_color="gray")
        self.lbl_info.pack(pady=5)
        
        # Lista de backups
        ctk.CTkLabel(self, text="Backups Disponibles:", 
                    font=("Arial", 13, "bold")).pack(anchor="w", padx=15, pady=(10, 5))
        
        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=10, pady=5)
        
        self._load_backups()
    
    def _crear_backup(self):
        self.lbl_info.configure(text="Creando backup...", text_color=Colors.WARNING)
        self.update()
        
        success, msg = self.backup_mgr.crear_backup()
        
        if success:
            messagebox.showinfo("Exito", f"Backup creado exitosamente:\n{Path(msg).name}")
            self.lbl_info.configure(text="Backup creado", text_color=Colors.SUCCESS)
            self._load_backups()
        else:
            messagebox.showerror("Error", f"Error al crear backup:\n{msg}")
            self.lbl_info.configure(text="Error al crear backup", text_color=Colors.DANGER)
    
    def _load_backups(self):
        for w in self.scroll.winfo_children():
            w.destroy()
        
        backups = self.backup_mgr.listar_backups()
        
        if not backups:
            ctk.CTkLabel(self.scroll, text="No hay backups disponibles", 
                       text_color="gray", font=("Arial", 12)).pack(pady=30)
            self.lbl_info.configure(text="0 backups encontrados")
            return
        
        self.lbl_info.configure(text=f"{len(backups)} backup(s) disponible(s)")
        
        for bk in backups:
            frame = ctk.CTkFrame(self.scroll, fg_color=Colors.CARD)
            frame.pack(fill="x", pady=3, padx=5)
            
            # Info frame
            info_f = ctk.CTkFrame(frame, fg_color="transparent")
            info_f.pack(side="left", fill="both", expand=True, padx=15, pady=10)
            
            ctk.CTkLabel(info_f, text=bk['nombre'], 
                       font=("Arial", 12, "bold"), anchor="w").pack(anchor="w")
            
            fecha_str = bk['fecha'].strftime('%d/%m/%Y %H:%M:%S')
            ctk.CTkLabel(info_f, text=f"{fecha_str} | {bk['tamaño_kb']:.2f} KB", 
                       font=("Arial", 10), text_color="gray", anchor="w").pack(anchor="w")
            
            # Botones
            btn_f = ctk.CTkFrame(frame, fg_color="transparent")
            btn_f.pack(side="right", padx=10)
            
            ctk.CTkButton(btn_f, text="RESTAURAR", width=100, height=30,
                         fg_color=Colors.WARNING,
                         command=lambda p=bk['path'], n=bk['nombre']: 
                         self._restaurar_backup(p, n)).pack(pady=2)
            
            ctk.CTkButton(btn_f, text="ELIMINAR", width=100, height=30,
                         fg_color=Colors.DANGER,
                         command=lambda p=bk['path'], n=bk['nombre']: 
                         self._eliminar_backup(p, n)).pack(pady=2)
    
    def _restaurar_backup(self, path: str, nombre: str):
        msg = (
            f"ADVERTENCIA\n\n"
            f"¿Desea restaurar el backup?\n\n"
            f"{nombre}\n\n"
            f"ESTO SOBRESCRIBIRA TODOS LOS DATOS ACTUALES\n"
            f"Se recomienda crear un backup antes de restaurar."
        )
        
        if not messagebox.askyesno("Confirmar Restauracion", msg):
            return
        
        self.lbl_info.configure(text="Restaurando backup...", text_color=Colors.WARNING)
        self.update()
        
        success, msg = self.backup_mgr.restaurar_backup(path)
        
        if success:
            messagebox.showinfo("Exito", "Backup restaurado exitosamente\n\nReinicie la aplicacion.")
            self.lbl_info.configure(text="Backup restaurado", text_color=Colors.SUCCESS)
            self.callback()
        else:
            messagebox.showerror("Error", f"Error al restaurar:\n{msg}")
            self.lbl_info.configure(text="Error al restaurar", text_color=Colors.DANGER)
    
    def _eliminar_backup(self, path: str, nombre: str):
        if not messagebox.askyesno("Confirmar", f"¿Eliminar backup?\n\n{nombre}"):
            return
        
        if self.backup_mgr.eliminar_backup(path):
            messagebox.showinfo("✅", "Backup eliminado")
            self._load_backups()
        else:
            messagebox.showerror("❌", "Error al eliminar backup")


# ============================================================================
# VENTANA RESET DE BD
# ============================================================================
class VentanaResetBD(ctk.CTkToplevel):
    def __init__(self, master, db: DBManager, backup_mgr: BackupManager, callback):
        super().__init__(master)
        self.title("RESETEAR BASE DE DATOS")
        self.geometry("600x550")
        self.attributes("-topmost", True)
        self.resizable(False, False)
        
        self.db = db
        self.backup_mgr = backup_mgr
        self.callback = callback
        
        # Header con advertencia
        header = ctk.CTkFrame(self, fg_color=Colors.DANGER)
        header.pack(fill="x")
        
        ctk.CTkLabel(header, text="RESETEAR BASE DE DATOS", 
                    font=("Arial", 20, "bold"), text_color="white").pack(pady=15)
        
        # Info
        info = ctk.CTkFrame(self, fg_color=Colors.CARD)
        info.pack(fill="both", expand=True, padx=20, pady=20)
        
        warning_text = (
            "ADVERTENCIA IMPORTANTE\n\n"
            "Esta operacion ELIMINARA PERMANENTEMENTE:\n\n"
            "- Todas las sesiones de inventario\n"
            "- Todos los items contados\n"
            "- Todo el historial de movimientos\n"
            "- Todos los equipos registrados\n\n"
            "SE RECOMIENDA CREAR UN BACKUP ANTES\n\n"
            "Esta accion NO se puede deshacer."
        )
        
        ctk.CTkLabel(info, text=warning_text, 
                    font=("Arial", 13), justify="left",
                    text_color=Colors.DANGER).pack(pady=20, padx=20)
        
        # Checkbox de confirmación
        self.var_confirmar = ctk.BooleanVar(value=False)
        self.cb_confirmar = ctk.CTkCheckBox(
            info, 
            text="Entiendo los riesgos y deseo continuar",
            variable=self.var_confirmar,
            font=("Arial", 12, "bold"),
            command=self._toggle_buttons
        )
        self.cb_confirmar.pack(pady=15)
        
        # Botones
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=15)
        
        self.btn_backup = ctk.CTkButton(
            btn_frame, 
            text="CREAR BACKUP PRIMERO",
            fg_color=Colors.SUCCESS, 
            height=45,
            font=("Arial", 13, "bold"),
            command=self._crear_backup_y_reset
        )
        self.btn_backup.pack(fill="x", pady=5)
        
        self.btn_reset = ctk.CTkButton(
            btn_frame, 
            text="RESETEAR SIN BACKUP",
            fg_color=Colors.DANGER, 
            height=45,
            font=("Arial", 13, "bold"),
            state="disabled",
            command=self._reset_bd
        )
        self.btn_reset.pack(fill="x", pady=5)
        
        ctk.CTkButton(
            btn_frame, 
            text="CANCELAR",
            fg_color="gray", 
            height=40,
            command=self.destroy
        ).pack(fill="x", pady=5)
    
    def _toggle_buttons(self):
        if self.var_confirmar.get():
            self.btn_reset.configure(state="normal")
        else:
            self.btn_reset.configure(state="disabled")
    
    def _crear_backup_y_reset(self):
        # Crear backup
        if messagebox.askyesno("Backup", "¿Crear backup antes de resetear?"):
            success, msg = self.backup_mgr.crear_backup()
            
            if not success:
                messagebox.showerror("Error", f"Error creando backup:\n{msg}")
                return
            
            messagebox.showinfo("✅", f"Backup creado:\n{Path(msg).name}")
        
        # Resetear
        self._reset_bd()
    
    def _reset_bd(self):
        if not messagebox.askyesno(
            "ÚLTIMA CONFIRMACIÓN",
            "¿Está COMPLETAMENTE SEGURO de resetear la base de datos?\n\nEsta acción es IRREVERSIBLE."
        ):
            return
        
        try:
            # Eliminar todas las tablas y recrearlas
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Desactivar foreign key checks
                cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
                
                # Eliminar tablas
                tablas = ['historial_movimientos', 'items_corte', 'equipos', 'sesiones']
                for tabla in tablas:
                    cursor.execute(f"DROP TABLE IF EXISTS {tabla}")
                    logger.info(f"Tabla {tabla} eliminada")
                
                # Reactivar foreign key checks
                cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
                
                conn.commit()
            
            # Recrear tablas
            self.db._create_tables()
            
            messagebox.showinfo(
                "Exito",
                "Base de datos reseteada exitosamente.\n\nLa aplicacion se reiniciara."
            )
            
            logger.info("Base de datos reseteada exitosamente")
            
            self.callback()
            self.destroy()
            
        except Exception as e:
            logger.error(f"Error reseteando BD: {e}")
            messagebox.showerror("Error", f"Error al resetear BD:\n{e}")


# ============================================================================
# VENTANA ACTUALIZAR STOCK
# ============================================================================
class VentanaActualizarStock(ctk.CTkToplevel):
    def __init__(self, master, db: DBManager, sesion_id: int, callback):
        super().__init__(master)
        self.title("ACTUALIZAR STOCK DEL CORTE")
        self.geometry("600x500")
        self.attributes("-topmost", True)
        self.resizable(False, False)
        
        self.db = db
        self.sesion_id = sesion_id
        self.callback = callback
        self.stock_calculado = {}
        
        # Header
        header = ctk.CTkFrame(self, fg_color="#6A1B9A")
        header.pack(fill="x")
        ctk.CTkLabel(header, text="ACTUALIZAR STOCK",
                    font=("Arial", 18, "bold"), text_color="white").pack(pady=15)
        
        # Info del corte
        info = ctk.CTkFrame(self, fg_color=Colors.CARD)
        info.pack(fill="x", padx=20, pady=10)
        
        try:
            sesion_info = self.db.execute_query(
                "SELECT nombre, fecha, responsable FROM sesiones WHERE id=%s",
                (self.sesion_id,), fetch=True
            )[0]
            
            ctk.CTkLabel(info, text=f"Corte: {sesion_info['nombre']}",
                        font=("Arial", 12, "bold")).pack(padx=15, pady=(10, 5))
            ctk.CTkLabel(info, text=f"Responsable: {sesion_info['responsable']}",
                        font=("Arial", 10)).pack(padx=15, pady=2)
        except Exception as e:
            logger.error(f"Error cargando info sesión: {e}")
        
        # Instrucciones
        instruc = ctk.CTkFrame(self, fg_color="#1F1F1F")
        instruc.pack(fill="x", padx=20, pady=10)
        
        instruc_text = (
            "INSTRUCCIONES:\n\n"
            "1. Seleccione el archivo Excel maestro\n"
            "2. El sistema recalculará el stock desde las hojas:\n"
            "   - CONDI, MAQUI, ASCINTEC\n"
            "3. Se actualizará SOLO el campo 'stock_sistema'\n"
            "4. Los conteos ya realizados NO se modificarán\n"
            "5. Las diferencias se recalcularán automáticamente"
        )
        
        ctk.CTkLabel(instruc, text=instruc_text, font=("Arial", 10),
                    justify="left", text_color="#FFD700").pack(padx=15, pady=15)
        
        # Botón cargar Excel
        ctk.CTkButton(self, text="CARGAR EXCEL Y ACTUALIZAR",
                     fg_color=Colors.SUCCESS, height=45, font=("Arial", 14, "bold"),
                     command=self._cargar_y_actualizar).pack(fill="x", padx=20, pady=20)
        
        # Info adicional
        self.lbl_info = ctk.CTkLabel(self, text="", font=("Arial", 11))
        self.lbl_info.pack(pady=10)
        
        ctk.CTkButton(self, text="CANCELAR", fg_color="gray",
                     height=35, command=self.destroy).pack(fill="x", padx=20, pady=(0, 20))
    
    def _cargar_y_actualizar(self):
        """Carga Excel y actualiza stock del corte"""
        # Forzar ventana al frente
        self.lift()
        self.focus_force()
        self.attributes('-topmost', True)
        self.update()
        
        path = filedialog.askopenfilename(
            title="Seleccionar Excel Maestro",
            filetypes=[("Excel", "*.xlsx *.xls")],
            parent=self
        )
        
        self.attributes('-topmost', False)
        self.lift()
        self.focus_force()
        
        if not path:
            return
        
        try:
            self.lbl_info.configure(text="Procesando Excel...", text_color=Colors.WARNING)
            self.update()
            
            # Calcular stock desde Excel
            self.stock_calculado = Utils.calcular_stock_desde_excel(path)
            
            if not self.stock_calculado:
                messagebox.showerror("Error", "No se pudo calcular stock del Excel")
                self.lbl_info.configure(text="Error al procesar", text_color=Colors.DANGER)
                return
            
            # Obtener items del corte
            items = self.db.execute_query(
                "SELECT id, codigo FROM items_corte WHERE sesion_id=%s",
                (self.sesion_id,), fetch=True
            )
            
            if not items:
                messagebox.showwarning("Advertencia", "El corte no tiene items")
                return
            
            # Confirmar actualización
            msg = (
                f"Se actualizará el stock de {len(items)} items\n\n"
                f"Stock calculado desde Excel: {len(self.stock_calculado)} códigos\n\n"
                f"¿Desea continuar?"
            )
            
            if not messagebox.askyesno("Confirmar Actualización", msg):
                self.lbl_info.configure(text="Operación cancelada", text_color="gray")
                return
            
            # Actualizar stock
            self.lbl_info.configure(text="Actualizando stock...", text_color=Colors.ACCENT)
            self.update()
            
            actualizados = 0
            sin_stock = 0
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                for item in items:
                    codigo = Utils.limpiar_codigo(item['codigo'])
                    stock = float(self.stock_calculado.get(codigo, 0.0))
                    
                    cursor.execute(
                        "UPDATE items_corte SET stock_sistema=%s WHERE id=%s",
                        (stock, item['id'])
                    )
                    
                    if stock > 0:
                        actualizados += 1
                    else:
                        sin_stock += 1
                
                conn.commit()
            
            logger.info(f"Stock actualizado: {actualizados} items con stock, {sin_stock} sin stock")
            
            # Mensaje de éxito
            messagebox.showinfo(
                "Actualización Exitosa",
                f"Stock actualizado correctamente\n\n"
                f"Items actualizados: {len(items)}\n"
                f"Con stock: {actualizados}\n"
                f"Sin stock: {sin_stock}"
            )
            
            self.lbl_info.configure(text="Actualización completada", text_color=Colors.SUCCESS)
            
            # Actualizar UI principal
            self.callback()
            
            # Cerrar ventana después de 2 segundos
            self.after(2000, self.destroy)
            
        except Exception as e:
            logger.error(f"Error actualizando stock: {e}")
            messagebox.showerror("Error", f"Error al actualizar stock:\n{e}")
            self.lbl_info.configure(text="Error", text_color=Colors.DANGER)


# ============================================================================
# VENTANA CONFIGURACIÓN
# ============================================================================
class VentanaConfiguracion(ctk.CTkToplevel):
    def __init__(self, master, config: dict, callback):
        super().__init__(master)
        self.title("CONFIGURACION DEL SISTEMA")
        self.geometry("500x400")
        self.attributes("-topmost", True)
        self.resizable(False, False)
        
        self.config = config
        self.callback = callback
        
        # Header
        header = ctk.CTkFrame(self, fg_color=Colors.INFO)
        header.pack(fill="x")
        ctk.CTkLabel(header, text="CONFIGURACION",
                    font=("Arial", 18, "bold"), text_color="white").pack(pady=15)
        
        # Formulario
        form = ctk.CTkFrame(self, fg_color=Colors.CARD)
        form.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Intervalo de sincronización
        ctk.CTkLabel(form, text="Intervalo de Sincronizacion (segundos):",
                    font=("Arial", 12, "bold")).pack(anchor="w", padx=20, pady=(20, 5))
        
        ctk.CTkLabel(form, text="Tiempo entre actualizaciones automáticas de KPIs y conteos",
                    font=("Arial", 10), text_color="gray").pack(anchor="w", padx=20)
        
        self.entry_interval = ctk.CTkEntry(form, height=35, font=("Arial", 14))
        self.entry_interval.pack(fill="x", padx=20, pady=10)
        self.entry_interval.insert(0, str(self.config['app'].get('sync_interval_seconds', 30)))
        self.entry_interval.bind("<Return>", lambda e: self._guardar())
        
        # Información adicional
        info_frame = ctk.CTkFrame(form, fg_color="#1F1F1F")
        info_frame.pack(fill="x", padx=20, pady=20)
        
        info_text = (
            "RECOMENDACIONES:\n\n"
            "- Minimo: 10 segundos\n"
            "- Recomendado: 30 segundos\n"
            "- Maximo: 120 segundos\n\n"
            "Valores muy bajos pueden afectar el rendimiento"
        )
        
        ctk.CTkLabel(info_frame, text=info_text, font=("Arial", 10),
                    justify="left", text_color="#FFD700").pack(padx=15, pady=15)
        
        # Botones
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=15)
        
        ctk.CTkButton(btn_frame, text="GUARDAR", fg_color=Colors.SUCCESS,
                     height=40, font=("Arial", 13, "bold"),
                     command=self._guardar).pack(side="left", fill="x", expand=True, padx=(0, 5))
        
        ctk.CTkButton(btn_frame, text="CANCELAR", fg_color="gray",
                     height=40, command=self.destroy).pack(side="right", fill="x", expand=True, padx=(5, 0))
    
    def _guardar(self):
        try:
            interval = int(self.entry_interval.get())
            
            if interval < 10:
                messagebox.showwarning("Validacion", "El intervalo minimo es 10 segundos")
                self.entry_interval.delete(0, 'end')
                self.entry_interval.insert(0, "30")
                return
            
            if interval > 120:
                messagebox.showwarning("Validacion", "El intervalo maximo es 120 segundos")
                self.entry_interval.delete(0, 'end')
                self.entry_interval.insert(0, "30")
                return
            
            # Actualizar configuración
            self.config['app']['sync_interval_seconds'] = interval
            
            # Guardar en archivo
            if ConfigManager.save(self.config):
                messagebox.showinfo("Exito", 
                    f"Configuracion guardada.\nIntervalo de sincronizacion: {interval} segundos\n\nSe aplicara en la proxima actualizacion.")
                self.callback(interval)  # Pasar nuevo intervalo
                self.destroy()
            else:
                messagebox.showerror("Error", "No se pudo guardar la configuracion")
        
        except ValueError:
            messagebox.showerror("Error", "Ingrese un numero valido")


# ============================================================================
# VENTANA EQUIPOS
# ============================================================================
class VentanaEquipos(ctk.CTkToplevel):
    def __init__(self, master, db: DBManager, callback):
        super().__init__(master)
        self.title("GESTIÓN DE EQUIPOS")
        self.geometry("500x600")
        self.attributes("-topmost", True)
        
        self.db = db
        self.callback = callback
        
        header = ctk.CTkFrame(self, fg_color=Colors.CARD)
        header.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(header, text="GESTION DE EQUIPOS",
                    font=("Arial", 18, "bold")).pack(pady=10)
        
        form = ctk.CTkFrame(self, fg_color=Colors.CARD)
        form.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(form, text="Numero de Equipo:").pack(anchor="w", padx=15, pady=(10, 5))
        
        self.entry_num = ctk.CTkEntry(form, placeholder_text="Ej: 1, 2, 3", height=35)
        self.entry_num.pack(fill="x", padx=15, pady=5)
        
        ctk.CTkLabel(form, text="Integrantes:").pack(anchor="w", padx=15, pady=(10, 5))
        
        self.entry_integrantes = ctk.CTkEntry(form, placeholder_text="Ej: JOSE MEJIA Y PRACTICANTE", height=35)
        self.entry_integrantes.pack(fill="x", padx=15, pady=5)
        self.entry_integrantes.bind("<Return>", lambda e: self._add_equipo())
        
        ctk.CTkButton(form, text="AGREGAR", fg_color=Colors.SUCCESS,
                     height=35, command=self._add_equipo).pack(fill="x", padx=15, pady=10)
        
        ctk.CTkLabel(self, text="Equipos Activos:",
                    font=("Arial", 12, "bold")).pack(anchor="w", padx=15, pady=(10, 5))
        
        self.scroll = ctk.CTkScrollableFrame(self)
        self.scroll.pack(fill="both", expand=True, padx=10, pady=5)
        
        self._load_equipos()
    
    def _load_equipos(self):
        for w in self.scroll.winfo_children():
            w.destroy()
        
        try:
            eqs = self.db.execute_query(
                "SELECT id, nombre_equipo, integrantes FROM equipos WHERE activo=1 ORDER BY nombre_equipo",
                fetch=True
            )
            
            if not eqs:
                ctk.CTkLabel(self.scroll, text="Sin equipos", 
                           text_color="gray").pack(pady=20)
                return
            
            for eq in eqs:
                frame = ctk.CTkFrame(self.scroll, fg_color="#333")
                frame.pack(fill="x", pady=3, padx=5)
                
                info_frame = ctk.CTkFrame(frame, fg_color="transparent")
                info_frame.pack(side="left", fill="both", expand=True, padx=15, pady=10)
                
                ctk.CTkLabel(info_frame, text=f"Equipo {eq['nombre_equipo']}", 
                           font=("Arial", 13, "bold"), anchor="w").pack(anchor="w")
                
                if eq.get('integrantes'):
                    ctk.CTkLabel(info_frame, text=eq['integrantes'], 
                               font=("Arial", 10), anchor="w",
                               text_color="gray").pack(anchor="w")
                
                ctk.CTkButton(frame, text="X", width=30, fg_color=Colors.DANGER,
                            command=lambda x=eq['id'], n=eq['nombre_equipo']: 
                            self._remove_equipo(x, n)).pack(side="right", padx=10)
        except Exception as e:
            logger.error(f"Error cargando equipos: {e}")
    
    def _add_equipo(self):
        numero = self.entry_num.get().strip()
        integrantes = self.entry_integrantes.get().strip()
        
        if not numero:
            messagebox.showwarning("Validación", "Ingrese el número del equipo")
            return
        
        if not integrantes:
            messagebox.showwarning("Validación", "Ingrese los integrantes del equipo")
            return
        
        integrantes = integrantes.upper()
        
        try:
            self.db.execute_query(
                "INSERT INTO equipos (nombre_equipo, integrantes) VALUES (%s, %s)",
                (numero, integrantes)
            )
            
            self.entry_num.delete(0, 'end')
            self.entry_integrantes.delete(0, 'end')
            self._load_equipos()
            self.callback()
            logger.info(f"Equipo agregado: {numero} - {integrantes}")
        except Error as e:
            if "Duplicate entry" in str(e):
                messagebox.showerror("Error", "Número de equipo ya existe")
            else:
                messagebox.showerror("Error", f"Error: {e}")
    
    def _remove_equipo(self, eq_id: int, nombre: str):
        if not messagebox.askyesno("Confirmar", f"¿Eliminar '{nombre}'?"):
            return
        
        try:
            self.db.execute_query(
                "UPDATE equipos SET activo=0 WHERE id=%s",
                (eq_id,)
            )
            self._load_equipos()
            self.callback()
            logger.info(f"Equipo eliminado: {nombre}")
        except Exception as e:
            messagebox.showerror("Error", f"Error: {e}")


# FIN PARTE 1
# Continúa en PARTE 2 con VentanaNuevoCorte y inicio de InventarioApp
"""
SISTEMA DE INVENTARIO V7.0 HÍBRIDO
===================================
PARTE 2 DE 3: VentanaNuevoCorte e Inicio de InventarioApp

Esta parte debe ir después de la PARTE 1
"""


# ============================================================================
# VENTANA NUEVO CORTE
# ============================================================================
class VentanaNuevoCorte(ctk.CTkToplevel):
    def __init__(self, master, db: DBManager, callback):
        super().__init__(master)
        self.title("NUEVO CORTE DE INVENTARIO")
        self.geometry("600x500")
        self.attributes("-topmost", True)
        self.resizable(False, False)
        
        self.db = db
        self.callback = callback
        self.df_maestro = None
        self.stock_calculado = {}
        
        header = ctk.CTkFrame(self, fg_color=Colors.ACCENT)
        header.pack(fill="x")
        
        ctk.CTkLabel(header, text="NUEVO CORTE", 
                    font=("Arial", 20, "bold"), text_color="white").pack(pady=15)
        
        form = ctk.CTkFrame(self, fg_color=Colors.CARD)
        form.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(form, text="Nombre del Corte:", 
                    font=("Arial", 12, "bold")).pack(anchor="w", padx=15, pady=(15, 5))
        self.en_nombre = ctk.CTkEntry(form, placeholder_text="Ej: INVENTARIO Q4 2024", height=35)
        self.en_nombre.pack(fill="x", padx=15, pady=5)
        
        ctk.CTkLabel(form, text="Responsable:", 
                    font=("Arial", 12, "bold")).pack(anchor="w", padx=15, pady=(15, 5))
        self.en_responsable = ctk.CTkEntry(form, placeholder_text="Nombre", height=35)
        self.en_responsable.pack(fill="x", padx=15, pady=5)
        
        ctk.CTkLabel(form, text="Bodega:", 
                    font=("Arial", 12, "bold")).pack(anchor="w", padx=15, pady=(15, 5))
        self.cmb_bodega = ctk.CTkComboBox(form, 
            values=["PRINCIPAL", "SECUNDARIA"], height=35)
        self.cmb_bodega.pack(fill="x", padx=15, pady=5)
        
        ctk.CTkButton(form, text="CARGAR EXCEL Y SELECCIONAR LINEAS", 
                     fg_color=Colors.SUCCESS, height=45, font=("Arial", 14, "bold"),
                     command=self._load_excel).pack(fill="x", padx=15, pady=20)
        
        self.lbl_info = ctk.CTkLabel(form, 
            text="Hojas requeridas:\nPRODUCTOS, CONDI, MAQUI, ASCINTEC",
            text_color="gray", font=("Arial", 10))
        self.lbl_info.pack(pady=10)
    
    def _load_excel(self):
        if not self.en_nombre.get().strip():
            self.lift()
            self.attributes('-topmost', True)
            self.after(100, lambda: self.attributes('-topmost', False))
            messagebox.showwarning("!", "Ingrese nombre del corte")
            return
        
        if not self.en_responsable.get().strip():
            self.lift()
            self.attributes('-topmost', True)
            self.after(100, lambda: self.attributes('-topmost', False))
            messagebox.showwarning("!", "Ingrese responsable")
            return
        
        # Forzar ventana al frente antes de abrir file dialog
        self.lift()
        self.focus_force()
        self.attributes('-topmost', True)
        self.update()
        
        path = filedialog.askopenfilename(
            title="Seleccionar Excel Maestro",
            filetypes=[("Excel", "*.xlsx *.xls")],
            parent=self
        )
        
        # Restaurar estado normal
        self.attributes('-topmost', False)
        self.lift()
        self.focus_force()
        
        if not path:
            return
        
        try:
            self.lbl_info.configure(text="Procesando...", text_color=Colors.WARNING)
            self.update()
            
            self.df_maestro = pd.read_excel(path, sheet_name="PRODUCTOS", dtype=str)
            self.df_maestro.columns = self.df_maestro.columns.str.strip().str.lower()
            
            self.stock_calculado = Utils.calcular_stock_desde_excel(path)
            
            # Cargar equipos desde Excel
            equipos_excel = Utils.cargar_equipos_desde_excel(path)
            if equipos_excel:
                self._agregar_equipos_desde_excel(equipos_excel)
            
            col_linea = 'deslinea' if 'deslinea' in self.df_maestro.columns else self.df_maestro.columns[2]
            lineas = sorted([str(x).strip() for x in self.df_maestro[col_linea].dropna().unique()])
            
            logger.info(f"Excel: {len(self.df_maestro)} productos, {len(lineas)} líneas")
            
            VentanaMultiSelect(self, lineas, self._save_corte)
        except Exception as e:
            logger.error(f"Error Excel: {e}")
            self.lift()
            self.attributes('-topmost', True)
            self.after(100, lambda: self.attributes('-topmost', False))
            messagebox.showerror("Error", f"Error:\n{e}")
            self.lbl_info.configure(text="Error", text_color=Colors.DANGER)
    
    def _agregar_equipos_desde_excel(self, equipos_data: List[dict]):
        """Agrega equipos desde Excel a la BD si no existen"""
        try:
            equipos_agregados = 0
            for equipo in equipos_data:
                try:
                    # Si es string (formato antiguo), solo agregar el número
                    if isinstance(equipo, str):
                        self.db.execute_query(
                            "INSERT INTO equipos (nombre_equipo) VALUES (%s)",
                            (equipo,)
                        )
                    # Si es dict con número e integrantes
                    elif isinstance(equipo, dict):
                        self.db.execute_query(
                            "INSERT INTO equipos (nombre_equipo, integrantes) VALUES (%s, %s)",
                            (equipo.get('numero', ''), equipo.get('integrantes', ''))
                        )
                    equipos_agregados += 1
                    logger.info(f"Equipo agregado desde Excel: {equipo}")
                except Error as e:
                    if "Duplicate entry" not in str(e):
                        logger.warning(f"Error agregando equipo {equipo}: {e}")
            
            if equipos_agregados > 0:
                logger.info(f"Total equipos agregados desde Excel: {equipos_agregados}")
        except Exception as e:
            logger.error(f"Error procesando equipos desde Excel: {e}")
    
    def _save_corte(self, lineas_sel: List[str]):
        if not lineas_sel:
            self.lift()
            self.attributes('-topmost', True)
            self.after(100, lambda: self.attributes('-topmost', False))
            messagebox.showwarning("!", "Seleccione al menos una línea")
            return
        
        try:
            sesion_id = self.db.execute_query(
                "INSERT INTO sesiones (nombre, fecha, responsable, bodega) VALUES (%s, %s, %s, %s)",
                (self.en_nombre.get().strip(), datetime.now(), 
                 self.en_responsable.get().strip(), self.cmb_bodega.get())
            )
            
            col_codigo = 'codproducto' if 'codproducto' in self.df_maestro.columns else self.df_maestro.columns[0]
            col_nombre = 'producto' if 'producto' in self.df_maestro.columns else self.df_maestro.columns[1]
            col_linea = 'deslinea' if 'deslinea' in self.df_maestro.columns else self.df_maestro.columns[2]
            
            items_data = []
            for _, row in self.df_maestro.iterrows():
                linea = str(row[col_linea]).strip()
                if linea in lineas_sel:
                    codigo = Utils.limpiar_codigo(row[col_codigo])
                    # Convertir explícitamente a float de Python
                    stock = float(self.stock_calculado.get(codigo, 0.0))
                    items_data.append((
                        sesion_id, codigo, str(row[col_nombre]).strip(), 
                        linea, stock
                    ))
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.executemany(
                    "INSERT INTO items_corte (sesion_id, codigo, producto, linea, stock_sistema) VALUES (%s,%s,%s,%s,%s)",
                    items_data
                )
                conn.commit()
            
            logger.info(f"Corte creado: ID={sesion_id}, Items={len(items_data)}")
            
            # Forzar ventana al frente para mensaje de éxito
            self.lift()
            self.attributes('-topmost', True)
            self.after(100, lambda: self.attributes('-topmost', False))
            messagebox.showinfo("Exito", 
                f"Corte creado\n\nItems: {len(items_data)}\nLíneas: {len(lineas_sel)}")
            
            self.callback(f"{sesion_id} - {self.en_nombre.get().strip()}")
            self.withdraw()
            self.after(100, self.destroy)
        except Exception as e:
            logger.error(f"Error creando corte: {e}")
            self.lift()
            self.attributes('-topmost', True)
            self.after(100, lambda: self.attributes('-topmost', False))
            messagebox.showerror("Error", f"Error:\n{e}")

# ============================================================================
# APLICACIÓN PRINCIPAL - PARTE 1: INIT Y SETUP
# ============================================================================
class InventarioApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("SISTEMA DE INVENTARIO - MySQL Multi-Usuario")
        self.geometry("1350x700")
        
        # Maximizar solo si la resolución lo permite
        try:
            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight()
            if screen_width >= 1400 and screen_height >= 900:
                self.state("zoomed")
        except Exception:
            pass
        
        # Inicializar DB
        try:
            self.db = DBManager()
            self.config = ConfigManager.load()
            self.backup_mgr = BackupManager(self.db)
        except Exception as e:
            logger.error(f"Error crítico: {e}")
            messagebox.showerror("Error Crítico", 
                f"No se pudo conectar a MySQL:\n{e}\n\nVerifique que MySQL esté corriendo.")
            sys.exit(1)
        
        # Variables de estado
        self.sesion_id = None
        self.equipo_id = None
        self.filtro_lineas = []
        self.producto_actual = {
            "codigo": None, "nombre": None, "linea": None, 
            "stock": 0.0, "es_nuevo": False
        }
        
        # Control de ventanas
        self.ventana_abierta = False
        self.console_visible = False
        
        # Control de sincronización
        self.sync_running = False
        self.sync_thread = None
        self.sync_interval_seconds = self.config['app'].get('sync_interval_seconds', 30)
        self.sync_counter = 0  # Contador para reducir actualizaciones pesadas
        self.sync_in_progress = False  # Flag para evitar sincronizaciones simultáneas
        
        # Data cache para tabs
        self.data_pendientes = []
        self.data_diferencias = []
        
        # Construir UI
        self._setup_ui()
        self._load_initial_data()
        
        # Iniciar sincronización automática
        self._start_sync()
        
        logger.info("App iniciada V7.0")
    
    def _setup_ui(self):
        """Configura interfaz responsive"""
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        self._create_sidebar()
        self._create_main_area()
        self._create_historial_panel()
        self._create_console()
    
    def _create_sidebar(self):
        """Panel lateral izquierdo"""
        sidebar_container = ctk.CTkFrame(self, width=220, corner_radius=0, fg_color=Colors.CARD)
        sidebar_container.grid(row=0, column=0, sticky="nsew")
        sidebar_container.grid_propagate(False)
        
        # Hacer el sidebar scrollable
        self.sidebar = ctk.CTkScrollableFrame(sidebar_container, fg_color=Colors.CARD)
        self.sidebar.pack(fill="both", expand=True)
        
        # Logo
        logo = ctk.CTkFrame(self.sidebar, fg_color=Colors.ACCENT)
        logo.pack(fill="x")
        ctk.CTkLabel(logo, text="INVENTARIO\nV7.0 HIBRIDO",
                    font=("Arial", 14, "bold"), text_color="white").pack(pady=10)
        
        # Sección Corte
        s1 = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        s1.pack(fill="x", padx=10, pady=(10, 5))
        
        ctk.CTkLabel(s1, text="1. CORTE:", font=("Arial", 10, "bold"),
                    anchor="w").pack(fill="x")
        self.cmb_sesion = ctk.CTkComboBox(s1, values=[], 
                                         command=self._on_select_sesion, height=28)
        self.cmb_sesion.pack(fill="x", pady=3)
        ctk.CTkButton(s1, text="+ NUEVO", fg_color=Colors.SUCCESS, 
                     height=25, command=self._new_corte).pack(fill="x", pady=2)
        
        # Sección Equipo
        s2 = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        s2.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(s2, text="2. EQUIPO:", font=("Arial", 10, "bold"),
                    anchor="w").pack(fill="x")
        self.cmb_equipo = ctk.CTkComboBox(s2, values=[], 
                                         command=self._on_select_equipo, height=28)
        self.cmb_equipo.pack(fill="x", pady=3)
        ctk.CTkButton(s2, text="GESTIONAR", fg_color="#555", 
                     height=25, command=self._manage_equipos).pack(fill="x", pady=2)
        
        # Sección Bodega
        s3 = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        s3.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(s3, text="3. BODEGA:", font=("Arial", 10, "bold"),
                    anchor="w").pack(fill="x")
        self.cmb_bodega = ctk.CTkComboBox(s3, 
            values=["PRINCIPAL", "SECUNDARIA"], height=28)
        self.cmb_bodega.set("PRINCIPAL")
        self.cmb_bodega.pack(fill="x", pady=3)
        
        # Estado
        self.lbl_status = ctk.CTkLabel(self.sidebar, text="SIN CONFIGURAR",
                                      font=("Arial", 11, "bold"), text_color=Colors.DANGER)
        self.lbl_status.pack(pady=10)
        
        # Indicador de sincronización
        self.lbl_sync = ctk.CTkLabel(self.sidebar, text="Sincronizando...", 
                                    font=("Arial", 9), text_color="gray")
        self.lbl_sync.pack(pady=3)
        
        # Botón de configuración
        ctk.CTkButton(self.sidebar, text="CONFIGURACION", fg_color="#555",
                     height=25, command=self._abrir_config).pack(fill="x", padx=10, pady=5)
        
        # Botones inferiores
        ctk.CTkButton(self.sidebar, text="LOG", fg_color="transparent",
                     border_width=1, height=25, command=self._toggle_console).pack(
                     side="bottom", fill="x", padx=10, pady=3)
        
        # Separador
        ctk.CTkFrame(self.sidebar, height=1, fg_color="gray").pack(
            side="bottom", fill="x", padx=10, pady=5)
        
        # Gestión de BD
        ctk.CTkLabel(self.sidebar, text="GESTION BD:", 
                    font=("Arial", 9, "bold"), text_color="gray").pack(
                    side="bottom", anchor="w", padx=10, pady=(5, 3))
        
        ctk.CTkButton(self.sidebar, text="BACKUPS", fg_color=Colors.INFO,
                     height=28, command=self._abrir_backups).pack(
                     side="bottom", fill="x", padx=10, pady=2)
        
        ctk.CTkButton(self.sidebar, text="RESETEAR BD", fg_color=Colors.DANGER,
                     height=28, command=self._abrir_reset).pack(
                     side="bottom", fill="x", padx=10, pady=2)
        
        ctk.CTkButton(self.sidebar, text="ACTUALIZAR STOCK", fg_color="#6A1B9A",
                     height=28, command=self._actualizar_stock_corte).pack(
                     side="bottom", fill="x", padx=10, pady=2)
        
        ctk.CTkButton(self.sidebar, text="EXPORTAR EXCEL", fg_color=Colors.SUCCESS, 
                     height=35, font=("Arial", 11, "bold"), 
                     command=self._export_excel).pack(side="bottom", fill="x", padx=10, pady=8)
    
    def _create_main_area(self):
        """Área principal central"""
        main_container = ctk.CTkFrame(self, fg_color="transparent")
        main_container.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        
        # Hacer el área principal scrollable
        self.main = ctk.CTkScrollableFrame(main_container, fg_color="transparent")
        self.main.pack(fill="both", expand=True)
        
        # Header con filtro
        header = ctk.CTkFrame(self.main, fg_color="transparent")
        header.pack(fill="x", pady=(0, 5))
        
        ctk.CTkButton(header, text="FILTRAR LINEAS", width=140, height=28,
                     command=self._open_filter).pack(side="left")
        
        self.lbl_filtro = ctk.CTkLabel(header, text="(Todas)", text_color="gray", 
                                      font=("Arial", 10))
        self.lbl_filtro.pack(side="left", padx=10)
        
        # KPIs
        self._create_kpi_cards()
        
        # Panel de operaciones
        self._create_operations_panel()
        
        # Tabs
        self._create_tabs()
    
    def _create_kpi_cards(self):
        """Tarjetas de KPIs"""
        kpi = ctk.CTkFrame(self.main, fg_color="transparent")
        kpi.pack(fill="x", pady=5)
        kpi.grid_columnconfigure((0,1,2,3,4,5), weight=1)
        
        self.kpi_avance = self._mk_kpi(kpi, "% AVANCE", "0%", Colors.ACCENT, 0)
        self.kpi_exactitud = self._mk_kpi(kpi, "EXACTITUD", "0%", Colors.SUCCESS, 1)
        self.kpi_pendientes = self._mk_kpi(kpi, "PENDIENTES", "0", Colors.INFO, 2)
        self.kpi_faltantes = self._mk_kpi(kpi, "FALTANTES", "0", Colors.DANGER, 3)
        self.kpi_sobrantes = self._mk_kpi(kpi, "SOBRANTES", "0", Colors.WARNING, 4)
        self.kpi_total = self._mk_kpi(kpi, "REGISTROS", "0", "gray", 5)
    
    def _mk_kpi(self, parent, title, value, color, col):
        """Crea tarjeta KPI individual"""
        card = ctk.CTkFrame(parent, fg_color=Colors.CARD)
        card.grid(row=0, column=col, sticky="ew", padx=1)
        
        ctk.CTkLabel(card, text=title, font=("Arial", 8), 
                    text_color="gray").pack(pady=(5, 1))
        
        lbl = ctk.CTkLabel(card, text=value, font=("Arial", 16, "bold"), 
                          text_color=color)
        lbl.pack(pady=(0, 5))
        return lbl
    
    def _create_operations_panel(self):
        """Panel de operaciones de conteo"""
        ops = ctk.CTkFrame(self.main, fg_color=Colors.CARD)
        ops.pack(fill="x", pady=5)
        
        # Info producto
        info = ctk.CTkFrame(ops, fg_color="#1F1F1F")
        info.pack(fill="x", padx=10, pady=8)
        
        # Instrucciones en la parte superior
        self.lbl_instrucciones = ctk.CTkLabel(info, 
            text="ESCANEE O ESCRIBA EL CODIGO > ENTER > INGRESE CANTIDAD > ENTER",
            font=("Arial", 9, "bold"), text_color="#00BFFF")
        self.lbl_instrucciones.pack(pady=3)
        
        self.lbl_producto = ctk.CTkLabel(info, text="ESPERANDO CÓDIGO...", 
                                        font=("Arial", 16, "bold"))
        self.lbl_producto.pack(pady=5)
        
        det_frame = ctk.CTkFrame(info, fg_color="transparent")
        det_frame.pack(pady=3)
        
        self.lbl_linea = ctk.CTkLabel(det_frame, text="", font=("Arial", 10))
        self.lbl_linea.pack(side="left", padx=10)
        
        self.lbl_stock = ctk.CTkLabel(det_frame, text="STOCK: ---", 
                                     font=("Arial", 14, "bold"), text_color=Colors.ACCENT)
        self.lbl_stock.pack(side="left", padx=10)
        
        # Inputs en una sola línea
        input_container = ctk.CTkFrame(ops, fg_color="transparent")
        input_container.pack(pady=5, fill="x", padx=10)
        
        # Fila 1: Código y Cantidad
        row1 = ctk.CTkFrame(input_container, fg_color="transparent")
        row1.pack(fill="x", pady=2)
        
        # Código
        codigo_frame = ctk.CTkFrame(row1, fg_color="transparent")
        codigo_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
        ctk.CTkLabel(codigo_frame, text="CODIGO:", 
                    font=("Arial", 10, "bold")).pack(anchor="w")
        self.en_codigo = ctk.CTkEntry(codigo_frame, height=35, 
                                     font=("Arial", 16))
        self.en_codigo.pack(fill="x", pady=2)
        self.en_codigo.bind("<Return>", lambda e: self._buscar_producto())
        
        # Cantidad
        cantidad_frame = ctk.CTkFrame(row1, fg_color="transparent")
        cantidad_frame.pack(side="left", padx=(5, 0))
        ctk.CTkLabel(cantidad_frame, text="CANTIDAD:", 
                    font=("Arial", 10, "bold")).pack(anchor="w")
        self.en_cantidad = ctk.CTkEntry(cantidad_frame, width=120, height=35, 
                                       justify="center", font=("Arial", 20, "bold"))
        self.en_cantidad.pack(pady=2)
        self.en_cantidad.insert(0, "1")
        self.en_cantidad.bind("<Return>", lambda e: self._pre_save())
        
        # Fila 2: Novedad
        row2 = ctk.CTkFrame(input_container, fg_color="transparent")
        row2.pack(fill="x", pady=(5, 0))
        ctk.CTkLabel(row2, text="NOVEDAD:", text_color="gray", 
                    font=("Arial", 9)).pack(anchor="w")
        self.en_novedad = ctk.CTkEntry(row2, height=25)
        self.en_novedad.pack(fill="x", pady=2)
        self.en_novedad.bind("<Return>", lambda e: self._pre_save())
        
        # Status
        self.lbl_op_status = ctk.CTkLabel(ops, text="Ingrese codigo y presione ENTER para buscar", 
                                         font=("Arial", 10, "bold"), text_color="#FFD700")
        self.lbl_op_status.pack(pady=3)
        
        # Info de conteo previo (se muestra cuando existe)
        self.lbl_conteo_previo = ctk.CTkLabel(ops, text="", 
                                             font=("Arial", 9), text_color="#FFA500")
        self.lbl_conteo_previo.pack(pady=2)
        
        self.en_codigo.focus()
    
    def _create_tabs(self):
        """Pestañas de información"""
        # Contenedor para tabs con altura máxima
        tabs_container = ctk.CTkFrame(self.main, fg_color="transparent")
        tabs_container.pack(fill="both", expand=True, pady=5)
        
        self.tabs = ctk.CTkTabview(tabs_container, height=180)
        self.tabs.pack(fill="both", expand=True)
        
        # Tab Búsqueda
        t1 = self.tabs.add("BUSQUEDA")
        self.en_busqueda = ctk.CTkEntry(t1, placeholder_text="Buscar...", height=30)
        self.en_busqueda.pack(fill="x", padx=10, pady=5)
        self.en_busqueda.bind("<KeyRelease>", lambda e: self._filtrar_busqueda())
        
        # Header de tabla
        header_busq = ctk.CTkFrame(t1, fg_color="#1F1F1F", height=30)
        header_busq.pack(fill="x", padx=10, pady=(5, 0))
        header_busq.pack_propagate(False)
        
        ctk.CTkLabel(header_busq, text="CÓDIGO", font=("Arial", 9, "bold"),
                    width=100, anchor="w").pack(side="left", padx=5)
        ctk.CTkLabel(header_busq, text="PRODUCTO", font=("Arial", 9, "bold"),
                    anchor="w").pack(side="left", fill="x", expand=True, padx=5)
        ctk.CTkLabel(header_busq, text="STOCK", font=("Arial", 9, "bold"),
                    width=80, anchor="center").pack(side="right", padx=(0, 5))
        ctk.CTkLabel(header_busq, text="ESTADO", font=("Arial", 9, "bold"),
                    width=180, anchor="center").pack(side="right", padx=5)
        
        self.scroll_busqueda = ctk.CTkScrollableFrame(t1)
        self.scroll_busqueda.pack(fill="both", expand=True, padx=10, pady=(0, 5))
        
        # Tab Pendientes
        t2 = self.tabs.add("PENDIENTES")
        
        # Barra superior con botón y búsqueda
        top_pend = ctk.CTkFrame(t2, fg_color="transparent")
        top_pend.pack(fill="x", padx=10, pady=5)
        ctk.CTkButton(top_pend, text="REFRESCAR", height=28, width=100,
                     command=self._load_pendientes).pack(side="left", padx=(0, 5))
        self.en_buscar_pend = ctk.CTkEntry(top_pend, placeholder_text="Filtrar...", height=28)
        self.en_buscar_pend.pack(side="left", fill="x", expand=True)
        self.en_buscar_pend.bind("<KeyRelease>", 
            lambda e: self._filtrar_precalc(self.en_buscar_pend.get(), 
                                           self.scroll_pendientes, self.data_pendientes))
        
        # Header de tabla
        header_pend = ctk.CTkFrame(t2, fg_color="#1F1F1F", height=30)
        header_pend.pack(fill="x", padx=10, pady=(5, 0))
        header_pend.pack_propagate(False)
        
        ctk.CTkLabel(header_pend, text="CÓDIGO", font=("Arial", 9, "bold"),
                    width=100, anchor="w").pack(side="left", padx=5)
        ctk.CTkLabel(header_pend, text="PRODUCTO", font=("Arial", 9, "bold"),
                    anchor="w").pack(side="left", fill="x", expand=True, padx=5)
        ctk.CTkLabel(header_pend, text="STOCK", font=("Arial", 9, "bold"),
                    width=80, anchor="center").pack(side="right", padx=5)
        
        self.scroll_pendientes = ctk.CTkScrollableFrame(t2)
        self.scroll_pendientes.pack(fill="both", expand=True, padx=10, pady=(0, 5))
        
        # Tab Diferencias
        t3 = self.tabs.add("DIFERENCIAS")
        
        # Barra superior con botón y búsqueda
        top_dif = ctk.CTkFrame(t3, fg_color="transparent")
        top_dif.pack(fill="x", padx=10, pady=5)
        ctk.CTkButton(top_dif, text="REFRESCAR", height=28, width=100,
                     command=self._load_diferencias).pack(side="left", padx=(0, 5))
        self.en_buscar_dif = ctk.CTkEntry(top_dif, placeholder_text="Filtrar...", height=28)
        self.en_buscar_dif.pack(side="left", fill="x", expand=True)
        self.en_buscar_dif.bind("<KeyRelease>", 
            lambda e: self._filtrar_precalc(self.en_buscar_dif.get(), 
                                           self.scroll_diferencias, self.data_diferencias))
        
        # Header de tabla
        header_dif = ctk.CTkFrame(t3, fg_color="#1F1F1F", height=30)
        header_dif.pack(fill="x", padx=10, pady=(5, 0))
        header_dif.pack_propagate(False)
        
        ctk.CTkLabel(header_dif, text="CÓDIGO", font=("Arial", 9, "bold"),
                    width=100, anchor="w").pack(side="left", padx=5)
        ctk.CTkLabel(header_dif, text="PRODUCTO", font=("Arial", 9, "bold"),
                    anchor="w").pack(side="left", fill="x", expand=True, padx=5)
        ctk.CTkLabel(header_dif, text="STOCK", font=("Arial", 9, "bold"),
                    width=80, anchor="center").pack(side="right", padx=(0, 5))
        ctk.CTkLabel(header_dif, text="DIFERENCIA", font=("Arial", 9, "bold"),
                    width=100, anchor="center").pack(side="right", padx=5)
        
        self.scroll_diferencias = ctk.CTkScrollableFrame(t3)
        self.scroll_diferencias.pack(fill="both", expand=True, padx=10, pady=(0, 5))
    
    def _create_historial_panel(self):
        """Panel derecho de historial"""
        hist_container = ctk.CTkFrame(self, width=200, corner_radius=0, fg_color="#232323")
        hist_container.grid(row=0, column=2, sticky="nsew")
        hist_container.grid_propagate(False)
        hist_container.grid_rowconfigure(1, weight=1)
        
        # Header fijo
        header_hist = ctk.CTkFrame(hist_container, fg_color="#232323")
        header_hist.grid(row=0, column=0, sticky="ew")
        ctk.CTkLabel(header_hist, text="HISTORIAL RECIENTE", 
                    font=("Arial", 11, "bold")).pack(pady=8)
        
        # Área scrollable
        self.hist_frame = ctk.CTkFrame(hist_container, fg_color="transparent")
        self.hist_frame.grid(row=1, column=0, sticky="nsew", padx=3, pady=3)
        
        self.scroll_historial = ctk.CTkScrollableFrame(self.hist_frame, fg_color="transparent")
        self.scroll_historial.pack(fill="both", expand=True)
    
    def _create_console(self):
        """Consola de logs"""
        self.console_frame = ctk.CTkFrame(self, height=120, corner_radius=0, fg_color=Colors.CONSOLE)
        self.console_frame.grid_propagate(False)
        
        self.console_box = ctk.CTkTextbox(self.console_frame, 
            font=("Consolas", 9), text_color="#00FF00", fg_color=Colors.CONSOLE)
        self.console_box.pack(fill="both", expand=True, padx=3, pady=3)
        self.console_box.configure(state="disabled")

    # ========================================================================
    # MÉTODOS DE CARGA Y CONFIGURACIÓN
    # ========================================================================
    
    def _load_initial_data(self):
        """Carga datos iniciales"""
        try:
            # Sesiones
            ses = self.db.execute_query(
                "SELECT id, nombre FROM sesiones WHERE activo=1 ORDER BY id DESC",
                fetch=True
            )
            self.cmb_sesion.configure(values=[f"{s['id']} - {s['nombre']}" for s in ses])
            
            # Equipos
            eqs = self.db.execute_query(
                "SELECT id, nombre_equipo, integrantes FROM equipos WHERE activo=1 ORDER BY nombre_equipo",
                fetch=True
            )
            # Crear diccionario con formato "Equipo X - NOMBRES"
            self.equipos_dict = {}
            equipos_display = []
            for e in eqs:
                integrantes = e.get('integrantes', '')
                if integrantes:
                    display = f"Equipo {e['nombre_equipo']} - {integrantes}"
                else:
                    display = f"Equipo {e['nombre_equipo']}"
                self.equipos_dict[display] = e['id']
                equipos_display.append(display)
            
            self.cmb_equipo.configure(values=equipos_display)
            if equipos_display:
                self.cmb_equipo.set(equipos_display[0])
                self.equipo_id = self.equipos_dict[equipos_display[0]]
            
            self._log("Datos iniciales cargados")
        except Exception as e:
            logger.error(f"Error cargando datos: {e}")
    
    def _new_corte(self):
        VentanaNuevoCorte(self, self.db, self._on_corte_created)
    
    def _on_corte_created(self, val):
        self._load_initial_data()
        self.cmb_sesion.set(val)
        self._on_select_sesion(val)
    
    def _manage_equipos(self):
        VentanaEquipos(self, self.db, self._load_initial_data)
    
    def _abrir_backups(self):
        """Abre ventana de gestión de backups"""
        VentanaBackups(self, self.backup_mgr, self._refresh_all)
    
    def _abrir_reset(self):
        """Abre ventana de reset de BD"""
        VentanaResetBD(self, self.db, self.backup_mgr, self._on_reset_complete)
    
    def _abrir_config(self):
        """Abre ventana de configuración"""
        VentanaConfiguracion(self, self.config, self._on_config_changed)
    
    def _actualizar_stock_corte(self):
        """Abre ventana para actualizar stock del corte actual"""
        if not self.sesion_id:
            messagebox.showwarning("Advertencia", "Seleccione un corte primero")
            return
        
        VentanaActualizarStock(self, self.db, self.sesion_id, self._refresh_all)
    
    def _on_reset_complete(self):
        """Callback después de resetear BD"""
        messagebox.showinfo("Reinicio", "Reiniciando aplicación...")
        self.destroy()
        # Reiniciar aplicación
        os.execl(sys.executable, sys.executable, *sys.argv)
    
    def _on_config_changed(self, new_interval=None):
        """Callback cuando se cambia la configuración"""
        self.config = ConfigManager.load()
        
        # Actualizar intervalo si se proporciona
        if new_interval is not None:
            self.sync_interval_seconds = new_interval
        else:
            self.sync_interval_seconds = self.config['app'].get('sync_interval_seconds', 30)
        
        self._log(f"Configuracion actualizada - Intervalo: {self.sync_interval_seconds}s")
    
    def _on_select_sesion(self, val):
        try:
            self.sesion_id = int(val.split(" - ")[0])
            self.filtro_lineas = []
            self._check_status()
            self._refresh_all()
            self._log(f"Sesión seleccionada: {self.sesion_id}")
        except Exception as e:
            logger.error(f"Error seleccionando sesión: {e}")
    
    def _on_select_equipo(self, val):
        self.equipo_id = self.equipos_dict.get(val)
        self._check_status()
        self._log(f"Equipo: {val}")
    
    def _check_status(self):
        if self.sesion_id and self.equipo_id:
            self.lbl_status.configure(text="SISTEMA LISTO", text_color=Colors.SUCCESS)
        else:
            self.lbl_status.configure(text="INCOMPLETO", text_color=Colors.DANGER)
    
    # ========================================================================
    # SINCRONIZACIÓN AUTOMÁTICA
    # ========================================================================
    
    def _start_sync(self):
        """Inicia sincronización en segundo plano"""
        self.sync_running = True
        self.sync_thread = threading.Thread(target=self._sync_loop, daemon=True)
        self.sync_thread.start()
        self._log("Sincronización automática iniciada")
    
    def _sync_loop(self):
        """Loop de sincronización automática"""
        while self.sync_running:
            try:
                time.sleep(self.sync_interval_seconds)
                
                if self.sesion_id:
                    # Actualizar en UI thread
                    self.after(0, self._sync_update)
            except Exception as e:
                logger.error(f"Error sync: {e}")
    
    def _sync_update(self):
        """Actualiza datos desde DB (iniciado desde UI thread, ejecuta en background)"""
        # Evitar múltiples sincronizaciones simultáneas
        if self.sync_in_progress:
            return
        
        self.sync_in_progress = True
        # Ejecutar actualización en hilo separado para no bloquear UI
        threading.Thread(target=self._sync_update_background, daemon=True).start()
    
    def _sync_update_background(self):
        """Ejecuta actualización en background thread"""
        if not self.sesion_id:
            return
            
        try:
            self.after(0, lambda: self.lbl_sync.configure(text="Actualizando..."))
            
            # === Consultas DB (en background thread) ===
            
            # 1. KPIs
            where = "WHERE sesion_id=%s"
            params = [self.sesion_id]
            
            if self.filtro_lineas:
                where += f" AND linea IN ({','.join(['%s']*len(self.filtro_lineas))})"
                params.extend(self.filtro_lineas)
            
            q_kpis = f"""SELECT COUNT(*) as total,
                    COUNT(CASE WHEN conteo_fisico>0 THEN 1 END) as contados,
                    COUNT(CASE WHEN (conteo_fisico-stock_sistema)<0 AND conteo_fisico>0 THEN 1 END) as faltantes,
                    COUNT(CASE WHEN (conteo_fisico-stock_sistema)>0 THEN 1 END) as sobrantes,
                    COUNT(CASE WHEN (conteo_fisico-stock_sistema)=0 AND conteo_fisico>0 THEN 1 END) as exactos
                FROM items_corte {where} AND stock_sistema > 0"""
            
            kpis = self.db.execute_query(q_kpis, tuple(params), fetch=True)[0]
            
            # 2. Pendientes
            q_pend = "SELECT codigo, producto, stock_sistema FROM items_corte WHERE sesion_id=%s AND conteo_fisico=0 AND stock_sistema > 0"
            p_pend = [self.sesion_id]
            if self.filtro_lineas:
                q_pend += f" AND linea IN ({','.join(['%s']*len(self.filtro_lineas))})"
                p_pend.extend(self.filtro_lineas)
            q_pend += " LIMIT 100"
            pends = self.db.execute_query(q_pend, tuple(p_pend), fetch=True)
            data_pend = [(x['codigo'], x['producto'], float(x['stock_sistema']) if x['stock_sistema'] else 0) for x in pends]
            
            # 3. Diferencias
            q_dif = """SELECT codigo, producto, diferencia, stock_sistema 
                       FROM items_corte WHERE sesion_id=%s AND diferencia!=0 AND conteo_fisico>0"""
            p_dif = [self.sesion_id]
            if self.filtro_lineas:
                q_dif += f" AND linea IN ({','.join(['%s']*len(self.filtro_lineas))})"
                p_dif.extend(self.filtro_lineas)
            q_dif += " LIMIT 100"
            difs = self.db.execute_query(q_dif, tuple(p_dif), fetch=True)
            data_dif = [(x['codigo'], x['producto'], x['diferencia'], float(x['stock_sistema']) if x['stock_sistema'] else 0) for x in difs]
            
            # 4. Historial
            movs = self.db.execute_query(
                """SELECT i.codigo, i.producto, i.conteo_fisico, i.diferencia, 
                          i.fecha_conteo, e.nombre_equipo
                   FROM items_corte i 
                   LEFT JOIN equipos e ON i.ultimo_equipo_id = e.id
                   WHERE i.sesion_id=%s AND i.conteo_fisico>0 
                   ORDER BY i.fecha_conteo DESC LIMIT 15""",
                (self.sesion_id,), fetch=True
            )
            
            # === Actualizar UI (desde main thread) ===
            
            # KPIs
            if kpis['total'] > 0:
                avance = (kpis['contados'] / kpis['total']) * 100
                self.after(0, lambda: self.kpi_avance.configure(text=f"{avance:.1f}%"))
                self.after(0, lambda: self.kpi_pendientes.configure(text=str(kpis['total'] - kpis['contados'])))
            
            if kpis['contados'] > 0:
                exactitud = (kpis['exactos'] / kpis['contados']) * 100
                self.after(0, lambda: self.kpi_exactitud.configure(text=f"{exactitud:.1f}%"))
            
            self.after(0, lambda: self.kpi_faltantes.configure(text=str(kpis['faltantes'])))
            self.after(0, lambda: self.kpi_sobrantes.configure(text=str(kpis['sobrantes'])))
            self.after(0, lambda: self.kpi_total.configure(text=str(kpis['total'])))
            
            # Tabs y Historial - solo actualizar cada 3 ciclos para reducir carga
            self.sync_counter += 1
            
            # Siempre actualizar cache de datos
            self.data_pendientes = data_pend
            self.data_diferencias = data_dif
            
            # Solo recrear widgets cada 3 sincronizaciones (90 segundos con intervalo de 30s)
            if self.sync_counter % 3 == 0:
                self.after(0, lambda: self._update_tabs_ui())
                self.after(0, lambda: self._update_historial_ui(movs))
            
            # Status
            self.after(0, lambda: self.lbl_sync.configure(
                text=f"Sincronizado {datetime.now().strftime('%H:%M:%S')}"
            ))
            
        except Exception as e:
            logger.error(f"Error sync_update: {e}")
            self.after(0, lambda: self.lbl_sync.configure(text="Error sync"))
        finally:
            # Liberar flag para permitir siguiente sincronización
            self.sync_in_progress = False
    
    # ========================================================================
    # FILTROS Y BÚSQUEDA
    # ========================================================================
    
    def _open_filter(self):
        if not self.sesion_id:
            messagebox.showwarning("!", "Seleccione un corte")
            return
        
        try:
            lins = self.db.execute_query(
                "SELECT DISTINCT linea FROM items_corte WHERE sesion_id=%s ORDER BY linea",
                (self.sesion_id,), fetch=True
            )
            VentanaMultiSelect(self, [l['linea'] for l in lins], self._apply_filter)
        except Exception as e:
            logger.error(f"Error filtro: {e}")
    
    def _apply_filter(self, sel):
        self.filtro_lineas = sel
        self.lbl_filtro.configure(
            text=f"{len(sel)} líneas" if sel else "(Todas)"
        )
        self._refresh_all()
        self._log(f"Filtro aplicado: {len(sel)} líneas")
    
    def _filtrar_busqueda(self):
        """Filtra búsqueda general en maestro"""
        if not self.sesion_id:
            return
        
        txt = self.en_busqueda.get().upper()
        
        for w in self.scroll_busqueda.winfo_children():
            w.destroy()
        
        if len(txt) < 2:
            return
        
        try:
            query = """SELECT codigo, producto, conteo_fisico, stock_sistema 
                       FROM items_corte 
                       WHERE sesion_id=%s AND (codigo LIKE %s OR producto LIKE %s)
                       LIMIT 30"""
            
            items = self.db.execute_query(
                query, (self.sesion_id, f"%{txt}%", f"%{txt}%"), fetch=True
            )
            
            for item in items:
                contado = float(item['conteo_fisico']) if item['conteo_fisico'] else 0.0
                stock = float(item['stock_sistema']) if item['stock_sistema'] else 0.0
                
                # Determinar estado: PENDIENTE, CONTADO o CUADRADO
                if contado == 0:
                    status = "PENDIENTE"
                    color = "#2B2B2B"  # Gris oscuro más tenue
                elif contado == stock:
                    status = f"CUADRADO: {contado:.0f}"
                    color = "#1F4D2E"  # Verde oscuro más tenue
                else:
                    dif = contado - stock
                    status = f"CONTADO: {contado:.0f} | Dif: {dif:+.0f}"
                    color = "#4D3D1F"  # Naranja oscuro más tenue
                
                self._mk_row_clickable_with_status(
                    self.scroll_busqueda, 
                    item['codigo'], 
                    item['producto'],
                    status,
                    color,
                    stock
                )
        except Exception as e:
            logger.error(f"Error búsqueda: {e}")
    
    def _update_tabs_ui(self):
        """Actualiza UI de tabs pendientes y diferencias (operación pesada)"""
        try:
            self._filtrar_precalc("", self.scroll_pendientes, self.data_pendientes)
            self._filtrar_precalc("", self.scroll_diferencias, self.data_diferencias)
        except Exception as e:
            logger.error(f"Error actualizando tabs: {e}")
    
    def _filtrar_precalc(self, txt, frame, data):
        """Filtra lista precalculada"""
        for w in frame.winfo_children():
            w.destroy()
        
        txt = txt.upper()
        count = 0
        
        for item in data:
            if count > 50:
                break
            
            cod = str(item[0])
            nom = str(item[1])
            
            if txt in cod or txt in nom.upper():
                # Extraer stock y diferencia si existen
                stock = item[2] if len(item) > 2 else None
                diferencia = item[3] if len(item) > 3 else None
                
                # Si tiene diferencia (viene de tab diferencias)
                if diferencia is not None:
                    self._mk_row_clickable(frame, cod, nom, stock, diferencia)
                # Si solo tiene stock (viene de tab pendientes)
                elif stock is not None:
                    self._mk_row_clickable(frame, cod, nom, stock, None)
                # Sin datos adicionales
                else:
                    self._mk_row_clickable(frame, cod, nom, None, None)
                count += 1
    
    def _mk_row_clickable(self, parent, codigo, texto, stock=None, diferencia=None):
        """Crea fila clickeable - estilo tabla con stock y diferencia opcionales"""
        frame = ctk.CTkFrame(parent, fg_color="#2B2B2B", height=32)
        frame.pack(fill="x", pady=1, padx=0)
        frame.pack_propagate(False)

        # Columna Código
        cod_frame = ctk.CTkFrame(frame, fg_color="transparent", width=100)
        cod_frame.pack(side="left", fill="y", padx=(5, 0))
        cod_frame.pack_propagate(False)
        
        ctk.CTkButton(
            cod_frame,
            text=codigo,
            anchor="w",
            fg_color="transparent",
            hover_color="#444",
            font=("Arial", 9),
            command=lambda: self._cargar_desde_lista(codigo)
        ).pack(fill="both", expand=True)

        # Columna Producto
        prod_frame = ctk.CTkFrame(frame, fg_color="transparent")
        prod_frame.pack(side="left", fill="both", expand=True, padx=5)
        
        ctk.CTkButton(
            prod_frame,
            text=texto[:45],
            anchor="w",
            fg_color="transparent",
            hover_color="#444",
            font=("Arial", 9),
            command=lambda: self._cargar_desde_lista(codigo)
        ).pack(fill="both", expand=True)

        # Columna Stock (si se proporciona)
        if stock is not None:
            stock_frame = ctk.CTkFrame(frame, fg_color="transparent", width=80)
            stock_frame.pack(side="right", fill="y", padx=(0, 5))
            stock_frame.pack_propagate(False)
            
            ctk.CTkLabel(
                stock_frame,
                text=f"{stock:.0f}",
                font=("Arial", 9),
                text_color="#FFFFFF"
            ).pack(expand=True)
        
        # Columna Diferencia (si se proporciona)
        if diferencia is not None:
            dif_frame = ctk.CTkFrame(frame, fg_color="transparent", width=100)
            dif_frame.pack(side="right", fill="y", padx=(0, 5))
            dif_frame.pack_propagate(False)
            
            dif_color = Colors.DANGER if diferencia < 0 else Colors.WARNING
            ctk.CTkLabel(
                dif_frame,
                text=f"{diferencia:+.0f}",
                font=("Arial", 9, "bold"),
                text_color=dif_color
            ).pack(expand=True)

    def _mk_row_clickable_with_status(self, parent, codigo, texto, status, color, stock=0):
        """Crea fila clickeable con estado de conteo - estilo tabla"""
        frame = ctk.CTkFrame(parent, fg_color=color, height=32)
        frame.pack(fill="x", pady=1, padx=0)
        frame.pack_propagate(False)

        # Columna Código
        cod_frame = ctk.CTkFrame(frame, fg_color="transparent", width=100)
        cod_frame.pack(side="left", fill="y", padx=(5, 0))
        cod_frame.pack_propagate(False)
        
        ctk.CTkButton(
            cod_frame,
            text=codigo,
            anchor="w",
            fg_color="transparent",
            hover_color="#555555",
            font=("Arial", 9),
            command=lambda: self._cargar_desde_lista(codigo)
        ).pack(fill="both", expand=True)

        # Columna Producto
        prod_frame = ctk.CTkFrame(frame, fg_color="transparent")
        prod_frame.pack(side="left", fill="both", expand=True, padx=5)
        
        ctk.CTkButton(
            prod_frame,
            text=texto[:35],
            anchor="w",
            fg_color="transparent",
            hover_color="#555555",
            font=("Arial", 9),
            command=lambda: self._cargar_desde_lista(codigo)
        ).pack(fill="both", expand=True)

        # Columna Stock
        stock_frame = ctk.CTkFrame(frame, fg_color="transparent", width=80)
        stock_frame.pack(side="right", fill="y", padx=(0, 5))
        stock_frame.pack_propagate(False)
        
        ctk.CTkLabel(
            stock_frame,
            text=f"{stock:.0f}",
            font=("Arial", 9),
            text_color="#FFFFFF"
        ).pack(expand=True)

        # Columna Estado
        if "CUADRADO" in status:
            status_color = "#4CAF50"  # Verde brillante
        elif "CONTADO" in status:
            status_color = "#FFA726"  # Naranja brillante
        else:
            status_color = "#FFD54F"  # Amarillo brillante para pendiente
        
        status_frame = ctk.CTkFrame(frame, fg_color="transparent", width=180)
        status_frame.pack(side="right", fill="y", padx=(0, 5))
        status_frame.pack_propagate(False)
        
        ctk.CTkLabel(
            status_frame,
            text=status,
            font=("Arial", 9, "bold"),
            text_color=status_color
        ).pack(expand=True)

    def _cargar_desde_lista(self, codigo):
        """Carga código desde lista"""
        self.en_codigo.delete(0, 'end')
        self.en_codigo.insert(0, codigo)
        self._buscar_producto()
    
    # ========================================================================
    # LÓGICA PRINCIPAL DE CONTEO
    # ========================================================================
    
    def _buscar_producto(self):
        """Busca producto por código"""
        if self.ventana_abierta:
            return
        
        if not self.sesion_id or not self.equipo_id:
            messagebox.showwarning("!", "Configure sesión y equipo")
            return
        
        cod = self.en_codigo.get()
        ok, res = InputValidator.validate_codigo(cod)
        
        if not ok:
            try:
                winsound.Beep(500, 300)  # Beep más grave para error
            except Exception:
                pass
            messagebox.showwarning("!", res)
            return
        
        cod = res
        self.en_codigo.delete(0, 'end')
        self.en_codigo.insert(0, cod)
        
        self.lbl_op_status.configure(text=f"Buscando codigo: {cod}...", text_color=Colors.ACCENT)
        self.lbl_conteo_previo.configure(text="")  # Limpiar mensaje previo
        self._log(f"Buscando: {cod}")
        
        try:
            prod = self.db.execute_query(
                """SELECT i.*, e.nombre_equipo FROM items_corte i 
                   LEFT JOIN equipos e ON i.ultimo_equipo_id=e.id 
                   WHERE i.sesion_id=%s AND i.codigo=%s""",
                (self.sesion_id, cod), fetch=True
            )
            
            if prod:
                p = prod[0]
                self.producto_actual = {
                    "codigo": cod,
                    "nombre": p['producto'],
                    "linea": p['linea'],
                    "stock": float(p['stock_sistema']),
                    "es_nuevo": False
                }
                
                self.lbl_producto.configure(text=p['producto'][:60], text_color="white")
                self.lbl_linea.configure(text=f"Línea: {p['linea']}")
                self.lbl_stock.configure(text=f"STOCK: {p['stock_sistema']}", 
                                        text_color=Colors.ACCENT)
                
                # Mostrar información de conteo previo si existe
                if p['conteo_fisico'] and p['conteo_fisico'] > 0:
                    fecha_conteo = ""
                    if p['fecha_conteo']:
                        fecha_conteo = p['fecha_conteo'].strftime("%d/%m %H:%M")
                    equipo_info = f"Equipo {p['nombre_equipo']}" if p['nombre_equipo'] else "Equipo desconocido"
                    self.lbl_conteo_previo.configure(
                        text=f"YA CONTADO: {p['conteo_fisico']} unidades por {equipo_info} el {fecha_conteo}"
                    )
                    
                    # Cargar novedad existente si hay
                    self.en_novedad.delete(0, 'end')
                    if p['novedad']:
                        self.en_novedad.insert(0, p['novedad'])
                else:
                    self.lbl_conteo_previo.configure(text="SIN CONTEO PREVIO - Primera vez")
                    # Limpiar novedad si es primera vez
                    self.en_novedad.delete(0, 'end')
                
                # Instrucción para ingresar cantidad
                self.lbl_op_status.configure(
                    text="Ingrese CANTIDAD y presione ENTER para guardar", 
                    text_color="#00FF00"
                )
                
                # DETECTAR CONFLICTO DE EQUIPO
                if p['conteo_fisico'] and p['conteo_fisico'] > 0:
                    if p['ultimo_equipo_id'] != self.equipo_id:
                        self._mostrar_conflicto_equipo(p)
                
                self.en_cantidad.delete(0, 'end')
                self.en_cantidad.insert(0, "1")
                self.en_cantidad.focus()
                self.en_cantidad.select_range(0, 'end')
                
            else:
                # Producto no existe
                try:
                    winsound.Beep(700, 400)  # Beep medio para advertencia
                except Exception:
                    pass
                
                self.lbl_op_status.configure(
                    text=f"ADVERTENCIA - Codigo '{cod}' NO ENCONTRADO", 
                    text_color=Colors.WARNING
                )
                
                if messagebox.askyesno("Nuevo", 
                    f"'{cod}' no existe.\n¿Agregar como EXTRA?"):
                    self._abrir_alta(cod)
                else:
                    self.lbl_op_status.configure(
                        text="Ingrese codigo y presione ENTER para buscar", 
                        text_color="#FFD700"
                    )
                    
        except Exception as e:
            logger.error(f"Error búsqueda: {e}")
            messagebox.showerror("Error", f"Error: {e}")
    
    def _mostrar_conflicto_equipo(self, producto):
        """Muestra alerta de conflicto entre equipos"""
        msg = (
            f"CONFLICTO DE EQUIPOS\n\n"
            f"Este producto ya fue contado por:\n"
            f"Equipo: {producto['nombre_equipo']}\n"
            f"Cantidad: {producto['conteo_fisico']}\n"
            f"Fecha: {producto['fecha_conteo']}\n\n"
            f"¿Desea modificar el conteo?"
        )
        
        respuesta = messagebox.askyesno("Conflicto de Equipo", msg)
        
        if not respuesta:
            self._limpiar_form()
    
    def _abrir_alta(self, cod):
        """Abre ventana para agregar producto nuevo"""
        self.ventana_abierta = True
        
        top = ctk.CTkToplevel(self)
        top.title("PRODUCTO NUEVO")
        top.geometry("400x300")
        top.attributes("-topmost", True)
        
        def close():
            self.ventana_abierta = False
            top.destroy()
            self.en_codigo.focus()
        
        top.protocol("WM_DELETE_WINDOW", close)
        
        ctk.CTkLabel(top, text=f"NUEVO: {cod}", 
                    font=("Arial", 16, "bold"), text_color=Colors.WARNING).pack(pady=10)
        
        ctk.CTkLabel(top, text="Nombre del Producto:").pack(anchor="w", padx=20)
        en_nom = ctk.CTkEntry(top, height=35)
        en_nom.pack(fill="x", padx=20, pady=5)
        en_nom.focus()
        
        ctk.CTkLabel(top, text="Línea:").pack(anchor="w", padx=20, pady=(10, 0))
        en_lin = ctk.CTkEntry(top, height=35)
        en_lin.pack(fill="x", padx=20, pady=5)
        
        def confirmar():
            nom = en_nom.get().strip().upper()
            lin = en_lin.get().strip().upper()
            
            if not nom:
                messagebox.showwarning("!", "Ingrese nombre")
                return
            
            # Agregar a BD
            try:
                self.db.execute_query(
                    """INSERT INTO items_corte (sesion_id, codigo, producto, linea, stock_sistema) 
                       VALUES (%s, %s, %s, %s, 0)""",
                    (self.sesion_id, cod, nom, lin or "SIN LINEA")
                )
                
                self.producto_actual = {
                    "codigo": cod,
                    "nombre": nom,
                    "linea": lin or "SIN LINEA",
                    "stock": 0.0,
                    "es_nuevo": True
                }
                
                self.lbl_producto.configure(text=f"NUEVO: {nom}", text_color=Colors.SUCCESS)
                self.lbl_stock.configure(text="STOCK: 0.00", text_color=Colors.WARNING)
                
                self.en_cantidad.focus()
                self.en_cantidad.select_range(0, 'end')
                
                self._log(f"Producto EXTRA agregado: {cod}")
                close()
                
            except Exception as e:
                messagebox.showerror("Error", f"Error: {e}")
        
        ctk.CTkButton(top, text="CONFIRMAR", fg_color=Colors.SUCCESS, 
                     height=40, command=confirmar).pack(pady=20)
    
    def _pre_save(self):
        """Prepara guardado"""
        if self.ventana_abierta:
            return
        
        cod = self.producto_actual["codigo"]
        if not cod:
            messagebox.showwarning("!", "Busque un producto primero")
            return
        
        cant_str = self.en_cantidad.get()
        ok, cant = InputValidator.validate_cantidad(cant_str)
        
        if not ok:
            messagebox.showwarning("!", cant)
            return
        
        try:
            # Verificar si ya existe conteo
            existe = self.db.execute_query(
                """SELECT ic.conteo_fisico, ic.ultimo_equipo_id, e.nombre_equipo, ic.fecha_conteo
                   FROM items_corte ic
                   LEFT JOIN equipos e ON ic.ultimo_equipo_id = e.id
                   WHERE ic.sesion_id=%s AND ic.codigo=%s""",
                (self.sesion_id, cod), fetch=True
            )
            
            if existe and existe[0]['conteo_fisico'] > 0:
                self._mostrar_dialogo_duplicado(cod, existe[0], cant)
            else:
                self._guardar_conteo(cod, cant, "NUEVO", 0)
                
        except Exception as e:
            logger.error(f"Error pre-save: {e}")
    
    def _mostrar_dialogo_duplicado(self, cod, item_data, cant_nueva):
        """Muestra diálogo para duplicados"""
        self.ventana_abierta = True
        # Convertir Decimal a float para evitar errores de tipo
        cant_actual = float(item_data['conteo_fisico'])
        cant_nueva = float(cant_nueva)
        equipo_nombre = item_data['nombre_equipo'] or "Desconocido"
        fecha_conteo = item_data['fecha_conteo']
        
        dlg = ctk.CTkToplevel(self)
        dlg.title("YA CONTADO")
        dlg.geometry("550x550")
        dlg.attributes("-topmost", True)
        dlg.resizable(False, False)
        
        def close():
            self.ventana_abierta = False
            dlg.destroy()
            self.en_codigo.focus()
        
        dlg.protocol("WM_DELETE_WINDOW", close)
        
        # Header
        header = ctk.CTkFrame(dlg, fg_color=Colors.WARNING)
        header.pack(fill="x")
        ctk.CTkLabel(header, text="YA CONTADO",
                     font=("Arial", 18, "bold"), text_color="white").pack(pady=15)

        # Información del conteo anterior
        info_frame = ctk.CTkFrame(dlg, fg_color=Colors.CARD)
        info_frame.pack(fill="x", padx=20, pady=15)

        ctk.CTkLabel(info_frame, text="CONTEO ANTERIOR:",
                     font=("Arial", 12, "bold")).pack(pady=(10, 5))

        ctk.CTkLabel(info_frame, text=f"Equipo: {equipo_nombre}",
                     font=("Arial", 11)).pack(pady=2)

        if fecha_conteo:
            fecha_str = fecha_conteo.strftime("%d/%m/%Y %H:%M")
            ctk.CTkLabel(info_frame, text=f"Fecha: {fecha_str}",
                         font=("Arial", 11)).pack(pady=2)

        ctk.CTkLabel(info_frame, text=f"Cantidad: {cant_actual:.2f}",
                     font=("Arial", 13, "bold"),
                     text_color=Colors.ACCENT).pack(pady=(5, 10))

        # Cantidad nueva
        ctk.CTkLabel(dlg, text=f"Cantidad Nueva: {cant_nueva:.2f}",
                     font=("Arial", 13, "bold")).pack(pady=10)
        
        # Descripción
        ctk.CTkLabel(dlg, text="Seleccione una acción:",
                     font=("Arial", 11), text_color="gray").pack(pady=(0, 10))
        
        # Botones de acción
        frame = ctk.CTkFrame(dlg, fg_color="transparent")
        frame.pack(pady=10, fill="x", padx=40)
        
        def sumar():
            self._guardar_conteo(cod, cant_actual + cant_nueva, "SUMA", cant_actual)
            close()
        
        def reemplazar():
            self._guardar_conteo(cod, cant_nueva, "REEMPLAZO", cant_actual)
            close()
        
        # Botón SUMAR (más grande y claro)
        btn_frame_sumar = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame_sumar.pack(fill="x", pady=5)
        
        ctk.CTkButton(
            btn_frame_sumar,
            text=f"SUMAR\n{cant_actual:.2f} + {cant_nueva:.2f} = {cant_actual + cant_nueva:.2f}",
            fg_color=Colors.SUCCESS,
            hover_color="#2d7a3e",
            height=70,
            font=("Arial", 14, "bold"),
            command=sumar
        ).pack(fill="x", padx=5)

        # Botón REEMPLAZAR (más grande y claro)
        btn_frame_reemplazar = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame_reemplazar.pack(fill="x", pady=5)
        
        ctk.CTkButton(
            btn_frame_reemplazar,
            text=f"REEMPLAZAR\n{cant_actual:.2f} -> {cant_nueva:.2f}",
            fg_color=Colors.ACCENT,
            hover_color="#1f538d",
            height=70,
            font=("Arial", 14, "bold"),
            command=reemplazar
        ).pack(fill="x", padx=5)
    
    def _guardar_conteo(self, cod, cant, tipo, cant_ant):
        """Guarda conteo en BD (ejecuta en background thread)"""
        nov = self.en_novedad.get().strip()
        if self.producto_actual["es_nuevo"]:
            nov = f"(NUEVO) {nov}"
        
        nombre_producto = self.producto_actual.get("nombre", cod)[:40]
        
        # Ejecutar guardado en background para no bloquear UI
        def guardar_background():
            try:
                # Actualizar item
                self.db.execute_query(
                    """UPDATE items_corte 
                       SET conteo_fisico=%s, novedad=%s, fecha_conteo=NOW(), ultimo_equipo_id=%s 
                       WHERE sesion_id=%s AND codigo=%s""",
                    (cant, nov, self.equipo_id, self.sesion_id, cod)
                )
                
                # Registrar historial
                self.db.execute_query(
                    """INSERT INTO historial_movimientos 
                       (sesion_id, item_codigo, equipo_id, tipo_accion, cantidad_anterior, cantidad_resultante, fecha_movimiento) 
                       VALUES (%s, %s, %s, %s, %s, %s, NOW())""",
                    (self.sesion_id, cod, self.equipo_id, tipo, cant_ant, cant)
                )
                
                # Actualizar UI desde main thread
                self.after(0, lambda: self._on_guardado_exitoso(nombre_producto, cant, tipo, cod))
                
            except Exception as e:
                logger.error(f"Error guardando: {e}")
                self.after(0, lambda: messagebox.showerror("Error", f"Error: {e}"))
        
        threading.Thread(target=guardar_background, daemon=True).start()
    
    def _on_guardado_exitoso(self, nombre_producto, cant, tipo, cod):
        """Callback cuando el guardado es exitoso (ejecuta en main thread)"""
        # BEEP de éxito
        try:
            winsound.Beep(1000, 200)
        except Exception:
            pass
        
        # Mensaje de éxito visual destacado
        mensaje_exito = f"OK - ITEM '{nombre_producto}' INVENTARIADO: {cant} unidades ({tipo})"
        self.lbl_op_status.configure(text=mensaje_exito, text_color=Colors.SUCCESS)
        self.lbl_conteo_previo.configure(text="")
        
        self._log(f"OK {tipo}: {cod} = {cant}")
        
        # Restaurar instrucción después de 3 segundos
        self.after(3000, lambda: self.lbl_op_status.configure(
            text="Ingrese codigo y presione ENTER para buscar", 
            text_color="#FFD700"
        ))
        
        # Limpiar y refrescar
        self._limpiar_form()
        self._refresh_all()
    
    def _limpiar_form(self):
        """Limpia formulario"""
        self.en_codigo.delete(0, 'end')
        self.en_cantidad.delete(0, 'end')
        self.en_cantidad.insert(0, "1")
        self.en_novedad.delete(0, 'end')
        self.lbl_producto.configure(text="ESPERANDO CODIGO...", text_color="white")
        self.lbl_linea.configure(text="")
        self.lbl_stock.configure(text="STOCK: ---")
        self.lbl_conteo_previo.configure(text="")
        self.producto_actual["codigo"] = None
        self.en_codigo.focus()
    
    # ========================================================================
    # ACTUALIZACIÓN DE DATOS
    # ========================================================================
    
    def _refresh_all(self):
        """Refresca todos los datos - forzar actualización completa"""
        self.sync_counter = 2  # Forzar actualización en el siguiente ciclo
        self._sync_update()  # Una sola llamada para todo
    
    def _update_kpis(self):
        """Actualiza KPIs - usa _sync_update para no duplicar código"""
        self._sync_update()
    
    def _load_historial(self):
        """Carga historial - usa _sync_update para no duplicar código"""
        self._sync_update()
    
    def _update_historial_ui(self, movs):
        """Actualiza UI del historial (se ejecuta en main thread)"""
        for w in self.scroll_historial.winfo_children():
            w.destroy()
        
        for m in movs:
            frame = ctk.CTkFrame(self.scroll_historial, fg_color="#2B2B2B")
            frame.pack(fill="x", pady=2, padx=3)
            
            # Producto
            ctk.CTkLabel(frame, text=m['producto'][:22], anchor="w",
                       font=("Arial", 10, "bold")).pack(fill="x", padx=5, pady=(5, 2))
            
            # Fila 1: Cantidad y Diferencia
            info_frame = ctk.CTkFrame(frame, fg_color="transparent")
            info_frame.pack(fill="x", padx=5, pady=1)
            
            ctk.CTkLabel(info_frame, text=f"Cant: {m['conteo_fisico']}", 
                       font=("Arial", 9, "bold")).pack(side="left")
            
            dif = float(m['diferencia'] or 0)
            col = Colors.DANGER if dif != 0 else Colors.SUCCESS
            ctk.CTkLabel(info_frame, text=f"Dif: {dif:+.2f}", 
                       text_color=col, font=("Arial", 9, "bold")).pack(side="right")
            
            # Fila 2: Equipo y Fecha/Hora
            equipo_fecha_frame = ctk.CTkFrame(frame, fg_color="transparent")
            equipo_fecha_frame.pack(fill="x", padx=5, pady=(1, 5))
            
            equipo_text = f"Equipo {m['nombre_equipo']}" if m.get('nombre_equipo') else "Equipo N/A"
            ctk.CTkLabel(equipo_fecha_frame, text=equipo_text, anchor="w",
                       font=("Arial", 8), text_color="#FFD700").pack(side="left")
            
            if m.get('fecha_conteo'):
                fecha_str = m['fecha_conteo'].strftime("%d/%m %H:%M")
            else:
                fecha_str = "--/-- --:--"
            
            ctk.CTkLabel(equipo_fecha_frame, text=fecha_str, anchor="e",
                       font=("Arial", 8), text_color="gray").pack(side="right", padx=(0, 8))
    
    def _load_pendientes(self):
        """Carga pendientes - usa _sync_update para no duplicar código"""
        self._sync_update()
    
    def _load_diferencias(self):
        """Carga diferencias - usa _sync_update para no duplicar código"""
        self._sync_update()
    
    # ========================================================================
    # UTILIDADES UI
    # ========================================================================
    
    def _toggle_console(self):
        """Muestra/oculta consola"""
        if self.console_visible:
            self.console_frame.grid_forget()
            self.console_visible = False
        else:
            self.console_frame.grid(row=1, column=1, sticky="ew")
            self.console_visible = True
    
    def _log(self, msg):
        """Agrega mensaje a consola"""
        try:
            ts = datetime.now().strftime("%H:%M:%S")
            self.console_box.configure(state="normal")
            self.console_box.insert("end", f"[{ts}] {msg}\n")
            self.console_box.see("end")
            self.console_box.configure(state="disabled")
        except Exception:
            pass
    
    # ========================================================================
    # EXPORTACIÓN A EXCEL
    # ========================================================================
    
    def _export_excel(self):
        """Exporta reporte completo a Excel"""
        if not self.sesion_id:
            messagebox.showwarning("!", "Seleccione un corte")
            return
        
        try:
            fn = f"REPORTE_{self.sesion_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            with self.db.get_connection() as conn:
                # Hoja 1: Conteo completo con equipo
                df_conteo = pd.read_sql(
                    f"""SELECT i.*, e.nombre_equipo as equipo_numero, e.integrantes as equipo_integrantes 
                        FROM items_corte i 
                        LEFT JOIN equipos e ON i.ultimo_equipo_id=e.id 
                        WHERE i.sesion_id={self.sesion_id}""",
                    conn
                )
                
                # Hoja 2: Solo diferencias con equipo
                df_dif = pd.read_sql(
                    f"""SELECT i.*, e.nombre_equipo as equipo_numero, e.integrantes as equipo_integrantes 
                        FROM items_corte i 
                        LEFT JOIN equipos e ON i.ultimo_equipo_id=e.id 
                        WHERE i.sesion_id={self.sesion_id} AND i.diferencia!=0 AND i.conteo_fisico>0""",
                    conn
                )
                
                # Hoja 3: Pendientes
                df_pend = pd.read_sql(
                    f"SELECT * FROM items_corte WHERE sesion_id={self.sesion_id} AND conteo_fisico=0",
                    conn
                )
                
                # Hoja 4: Historial con equipo
                df_hist = pd.read_sql(
                    f"""SELECT h.*, e.nombre_equipo as equipo_numero, e.integrantes as equipo_integrantes 
                        FROM historial_movimientos h 
                        LEFT JOIN equipos e ON h.equipo_id=e.id 
                        WHERE h.sesion_id={self.sesion_id}""",
                    conn
                )
                
                with pd.ExcelWriter(fn, engine='openpyxl') as writer:
                    df_conteo.to_excel(writer, sheet_name="CONTEO_COMPLETO", index=False)
                    df_dif.to_excel(writer, sheet_name="DIFERENCIAS", index=False)
                    df_pend.to_excel(writer, sheet_name="PENDIENTES", index=False)
                    df_hist.to_excel(writer, sheet_name="HISTORIAL", index=False)
                
                # Pintar diferencias
                self._pintar_diferencias_excel(fn)
            
            messagebox.showinfo("✅", f"Exportado:\n{fn}")
            self._log(f"Exportado: {fn}")
            
        except Exception as e:
            logger.error(f"Error export: {e}")
            messagebox.showerror("Error", f"Error:\n{e}")
    
    def _pintar_diferencias_excel(self, filename):
        """Pinta celdas con diferencias"""
        try:
            wb = load_workbook(filename)
            
            if "DIFERENCIAS" in wb.sheetnames:
                ws = wb["DIFERENCIAS"]
                red = PatternFill(start_color="FF9999", fill_type="solid")
                
                headers = [cell.value for cell in ws[1]]
                if 'diferencia' in headers:
                    idx_dif = headers.index('diferencia') + 1
                    
                    for row in ws.iter_rows(min_row=2):
                        try:
                            if float(row[idx_dif - 1].value or 0) != 0:
                                for cell in row:
                                    cell.fill = red
                        except (ValueError, TypeError, AttributeError):
                            pass
            
            wb.save(filename)
        except Exception as e:
            logger.error(f"Error pintando Excel: {e}")
    
    # ========================================================================
    # CIERRE DE APLICACIÓN
    # ========================================================================
    
    def destroy(self):
        """Cierre controlado de la aplicación"""
        try:
            # Detener sincronización
            self.sync_running = False
            if self.sync_thread and self.sync_thread.is_alive():
                self.sync_thread.join(timeout=2)
            
            logger.info("Aplicación cerrada correctamente")
        except Exception as e:
            logger.error(f"Error al cerrar: {e}")
        finally:
            super().destroy()

# ============================================================================
# PUNTO DE ENTRADA
# ============================================================================
if __name__ == "__main__":
    try:
        app = InventarioApp()
        app.mainloop()
    except Exception as e:
        logger.error(f"Error crítico en main: {e}")
        messagebox.showerror("Error Crítico", f"Error al iniciar aplicación:\n{e}")
        sys.exit(1)